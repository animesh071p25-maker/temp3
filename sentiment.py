import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from textblob import TextBlob
import os
import re
import datetime
import requests
import io
from typing import Optional, Dict, Any

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import pytesseract
except Exception:
    pytesseract = None

# ── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Social Media Analytics",
    layout="wide",
    page_icon="",
)

# ── DYNAMIC CSS INHERITING STREAMLIT NATIVE THEME ─────────────────────────────
_BG       = "var(--background-color)"
_BG2      = "var(--secondary-background-color)"
_TEXT     = "var(--text-color)"
_BORDER   = "rgba(150, 150, 150, 0.3)"
_ACCENT   = "#6c63ff"
_ACCENT2  = "#4ecdc4"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] {{
    font-family: 'Inter', sans-serif;
}}
/* Menus / tabs / containers */
.stTabs [data-baseweb="tab-list"] {{
    gap: 8px;
    background: transparent;
    padding: 8px 8px 0 8px;
}}
.stTabs [data-baseweb="tab"] {{
    background: {_BG2};
    border-radius: 8px 8px 0 0;
    padding: 10px 20px;
    color: {_TEXT};
    font-weight: 500;
    border: 1px solid {_BORDER};
    border-bottom: none;
}}
.stTabs [aria-selected="true"] {{
    background: linear-gradient(135deg, {_ACCENT}, {_ACCENT2}) !important;
    color: white !important;
    border-color: transparent !important;
}}
[data-testid="metric-container"] {{
    background: linear-gradient(135deg, {_BG2} 0%, {_BG} 100%);
    border: 1px solid {_BORDER};
    border-radius: 12px;
    padding: 16px;
}}
.trend-card {{
    background: linear-gradient(135deg, {_BG2} 0%, {_BG} 100%);
    border: 1px solid {_BORDER};
    border-radius: 14px;
    padding: 18px 20px;
    margin: 8px 0;
    transition: border-color 0.2s;
}}
.trend-card:hover {{ border-color: {_ACCENT}; }}
.trend-title {{ color: {_TEXT}; font-size: 15px; font-weight: 600; margin-bottom: 4px; }}
.trend-meta  {{ opacity: 0.8; font-size: 12px; }}
.trend-source-badge {{
    display: inline-block; padding: 2px 10px; border-radius: 20px;
    font-size: 11px; font-weight: 600; margin-right: 6px;
}}
.badge-gnews  {{ background: {_ACCENT2}; color: #0a0a15; }}
.badge-trends {{ background: {_ACCENT}; color: #fff; }}
.hashtag-pill {{
    display: inline-block;
    background: {_BG2};
    border: 1px solid {_ACCENT}55;
    border-radius: 20px; padding: 4px 14px; margin: 4px 4px 4px 0;
    font-size: 13px; color: {_TEXT}; font-weight: 500;
    cursor: pointer; transition: all 0.2s;
}}
.hashtag-pill:hover {{ border-color: {_ACCENT}; }}
.section-header {{
    font-size: 20px; font-weight: 700; color: {_TEXT};
    border-left: 4px solid {_ACCENT}; padding-left: 12px; margin: 24px 0 16px 0;
}}
.rank-number {{
    display: inline-block; width: 28px; height: 28px; border-radius: 50%;
    background: linear-gradient(135deg, {_ACCENT}, {_ACCENT2});
    color: white; font-weight: 700; text-align: center;
    line-height: 28px; font-size: 13px;
}}
.post-card {{
    background: {_BG2}; border: 1px solid {_BORDER};
    border-radius: 12px; padding: 16px; margin: 10px 0;
}}
.post-caption {{ color: {_TEXT}; font-size: 14px; line-height: 1.6; }}
.post-stats   {{ opacity: 0.8; font-size: 12px; margin-top: 8px; }}
/* Sidebar hidden */
[data-testid="stSidebar"] {{ display: none !important; }}
section[data-testid="stSidebarContent"] {{ display: none !important; }}
</style>
""", unsafe_allow_html=True)


# ── GLOBAL MEDICAL KEYWORDS ───────────────────────────────────────────────────
MEDICAL_TOPICS = [
    "NGS", "Next Generation Sequencing", "Diagnostics", "Oncology", "Genomics",
    "Pathology", "Biomarker", "Infectious Disease", "Tumor Profiling", "Healthcare AI",
    "Liquid Biopsy", "Precision Medicine", "Molecular Diagnostics", "Clinical Genomics",
    "Sequencing", "Cancer Genomics", "Pharmacogenomics", "Rare Disease",
    "Genetic Testing", "Clinical Pathology",
]

# ── HASHTAG BANK ──────────────────────────────────────────────────────────────
HASHTAG_BANK = {
    "NGS":                     ["#NGS", "#NextGenSequencing", "#GenomicsNow", "#DNASequencing", "#WGS"],
    "Oncology":                ["#CancerResearch", "#Oncology", "#CancerGenomics", "#TumorBoard", "#CancerCare"],
    "Diagnostics":             ["#Diagnostics", "#ClinicalDiagnostics", "#LabMedicine", "#PathologyLife", "#DiagnosticMedicine"],
    "Genomics":                ["#Genomics", "#HumanGenome", "#GenomicMedicine", "#FunctionalGenomics", "#Omics"],
    "Pathology":               ["#Pathology", "#Histopathology", "#DigitalPathology", "#ClinicalPathology", "#PathologyLab"],
    "Biomarker":               ["#Biomarker", "#Biomarkers", "#PredictiveMedicine", "#BiomarkerDiscovery"],
    "Infectious Disease":      ["#InfectiousDisease", "#AMR", "#PathogenGenomics", "#InfectionControl"],
    "Liquid Biopsy":           ["#LiquidBiopsy", "#ctDNA", "#cfDNA", "#CancerEarlyDetection"],
    "Precision Medicine":      ["#PrecisionMedicine", "#PersonalizedMedicine", "#GenomicMedicine"],
    "Molecular Diagnostics":   ["#MolecularDiagnostics", "#MolDx", "#PCR", "#qPCR", "#NAAT"],
    "AI":                      ["#AIinHealthcare", "#MedicalAI", "#HealthTech", "#ClinicalAI", "#DiagnosticAI"],
    "General":                 ["#MedicalScience", "#HealthcareInnovation", "#ClinicalResearch",
                                 "#MedTwitter", "#LabLife", "#PathologyNation", "#MolecularMedicine"],
}

# ── DATA LOADING ──────────────────────────────────────────────────────────────
@st.cache_data
def load_account_metrics():
    """
    Load account-level daily metrics (followers gained, profile visits, etc.)
    from any CSV/XLSX in the app folder.

    Expected (flexible) columns in the file:
      - Date: any parseable date (DD/MM/YYYY, YYYY-MM-DD, etc.)
      - Account: account handle/name (e.g., identifi.health, HaystackAnalytics)
      - Platform (optional): Instagram/LinkedIn; if missing defaults to Instagram
      - Followers Gained: numeric (optional)
      - Profile Visits: numeric (optional)

    The function is intentionally permissive: it tries multiple column name aliases.
    """
    # Look only in the same directory as this script / working dir.
    candidates = []
    for fn in os.listdir("."):
        lf = fn.lower()
        if not (lf.endswith(".csv") or lf.endswith(".xlsx") or lf.endswith(".xls")):
            continue
        # Heuristic: likely metrics files
        if any(k in lf for k in ["follower", "followers", "profile visit", "profile_visit", "insights", "metrics", "identifi", "haystack", "insta", "instagram", "linkedin"]):
            candidates.append(fn)

    if not candidates:
        return pd.DataFrame(columns=["Date", "Account", "Platform", "Followers Gained", "Profile Visits"])

    def _read_any(path: str) -> pd.DataFrame:
        try:
            if path.lower().endswith(".csv"):
                return pd.read_csv(path, encoding="utf-8-sig", low_memory=False)
            # Excel
            wb = pd.read_excel(path, sheet_name=None)
            if isinstance(wb, dict):
                frames = [df for df in wb.values() if isinstance(df, pd.DataFrame) and not df.empty]
                return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
            return wb
        except Exception:
            return pd.DataFrame()

    def _pick_col(df: pd.DataFrame, aliases):
        cols = {c.strip().lower(): c for c in df.columns}
        for a in aliases:
            key = a.strip().lower()
            if key in cols:
                return cols[key]
        # fuzzy contains match (e.g. "followers gained (total)")
        for c in df.columns:
            cl = c.strip().lower()
            if any(key in cl for key in [a.strip().lower() for a in aliases]):
                return c
        return None

    frames = []
    for fn in candidates:
        raw = _read_any(fn)
        if raw.empty:
            continue

        date_col = _pick_col(raw, ["date", "day"])
        acct_col = _pick_col(raw, ["account", "account name", "username", "profile", "page"])
        link_col = _pick_col(raw, ["post link/ date", "post link/date", "post link", "post link/ date "])
        plat_col = _pick_col(raw, ["platform", "source", "channel"])
        fol_col  = _pick_col(raw, ["followers gained", "follower gained", "followers gain", "followers_gained", "follows", "new followers", "followers"])
        pv_col   = _pick_col(raw, ["profile visits", "profile visit", "profile_visits", "profile visit by account", "profile visits by account"])

        if not date_col:
            continue

        df = pd.DataFrame()
        df["Date"] = pd.to_datetime(raw[date_col].astype(str), errors="coerce").dt.date

        # Account inference:
        # - Preferred: explicit Account column
        # - Fallback: parse first token from "Post link/ date" style column (e.g. "identifi.health\nSep 27, 10:59")
        # - Final fallback: infer from filename for known accounts
        if acct_col and acct_col in raw.columns:
            df["Account"] = raw[acct_col].astype(str).str.strip()
        elif link_col and link_col in raw.columns:
            s = raw[link_col].astype(str)
            # pandas .str has no splitlines(); split on newline explicitly
            first_line = s.str.split(r"\r?\n", regex=True).str[0].fillna("")
            acct = first_line.str.extract(r"^\s*([A-Za-z0-9._]+)", expand=False).fillna("").str.strip()
            # Clean cases like "HaystackAnalyticsJan" or "identifi.healthSep" → strip trailing month token
            acct = acct.str.replace(
                r"(?i)(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)$",
                "",
                regex=True,
            ).str.strip()
            df["Account"] = acct
        else:
            lf = fn.lower()
            if "identifi" in lf:
                df["Account"] = "identifi.health"
            elif "haystack" in lf:
                # LinkedIn account uses a different label in the UI
                df["Account"] = "HaystackAnalytics" if "linkedin" in lf else "haystack_analytics"
            elif "infexn" in lf:
                df["Account"] = "infexn_in"
            else:
                df["Account"] = ""

        if plat_col and plat_col in raw.columns:
            df["Platform"] = raw[plat_col].astype(str).str.strip()
        else:
            # Infer platform from filename if possible; otherwise default to Instagram
            lf = fn.lower()
            df["Platform"] = "LinkedIn" if "linkedin" in lf else "Instagram"

        def _to_num_series(s):
            return pd.to_numeric(s.astype(str).str.replace(",", "", regex=False), errors="coerce")

        df["Followers Gained"] = _to_num_series(raw[fol_col]) if fol_col and fol_col in raw.columns else pd.NA
        df["Profile Visits"]   = _to_num_series(raw[pv_col])  if pv_col and pv_col in raw.columns else pd.NA

        # Drop rows without an inferred account
        df["Account"] = df["Account"].astype(str).str.strip()
        df = df.dropna(subset=["Date"])
        df = df[df["Account"] != ""]
        if df.empty:
            continue

        frames.append(df)

    if not frames:
        return pd.DataFrame(columns=["Date", "Account", "Platform", "Followers Gained", "Profile Visits"])

    out = pd.concat(frames, ignore_index=True)
    # Normalize platform labels a bit
    out["Platform"] = out["Platform"].replace(
        {
            "ig": "Instagram",
            "insta": "Instagram",
            "instagram": "Instagram",
            "li": "LinkedIn",
            "linkedin": "LinkedIn",
        }
    )
    # If duplicates exist across files, sum by day/account/platform
    out = out.groupby(["Date", "Account", "Platform"], as_index=False).agg(
        {"Followers Gained": "sum", "Profile Visits": "sum"}
    )
    return out


@st.cache_data
def load_profile_visit_images() -> pd.DataFrame:
    """
    Read profile visit screenshots (e.g., 'haystack insta.jpeg') and extract numeric metrics.

    - If OCR is available (pytesseract + Pillow + local tesseract install), it will be used.
    - Otherwise, we return an empty df and the UI will offer manual entry fallback.

    Returns columns:
      - Account, Platform, Profile Visits, External Link Taps, SourceFile
    """
    img_exts = (".png", ".jpg", ".jpeg", ".webp")
    imgs = []
    for fn in os.listdir("."):
        lf = fn.lower().strip()
        if lf.endswith(img_exts):
            # heuristic: only files that look like account insights screenshots
            if any(k in lf for k in ["profile", "visit", "visits", "identifi", "haystack", "infexn", "insta", "instagram", "linkedin"]):
                imgs.append(fn)

    if not imgs:
        return pd.DataFrame(columns=["Account", "Platform", "Profile Visits", "External Link Taps", "SourceFile"])

    if Image is None or pytesseract is None:
        return pd.DataFrame(columns=["Account", "Platform", "Profile Visits", "External Link Taps", "SourceFile"])

    def _infer_account_platform(filename: str) -> tuple[Optional[str], Optional[str]]:
        lf = filename.lower()
        platform = None
        if "linkedin" in lf or "li" in lf:
            platform = "LinkedIn"
        if "insta" in lf or "instagram" in lf:
            platform = "Instagram"

        # map using your configured accounts
        if "haystack" in lf:
            return ("haystack_analytics" if platform == "Instagram" else "HaystackAnalytics", platform or "Instagram")
        if "identifi" in lf:
            return ("identifi.health", platform or "Instagram")
        if "infexn" in lf:
            return ("infexn_in", platform or "Instagram")
        return (None, platform)

    def _extract_int_after_label(text: str, label_regex: str) -> Optional[int]:
        m = re.search(label_regex, text, flags=re.IGNORECASE)
        if not m:
            return None
        tail = text[m.end(): m.end() + 40]
        m2 = re.search(r"(\d[\d,\. ]{0,15})", tail)
        if not m2:
            return None
        num = re.sub(r"[^\d]", "", m2.group(1))
        return int(num) if num else None

    rows = []
    for fn in imgs:
        acct, plat = _infer_account_platform(fn)
        try:
            img = Image.open(fn)
            # upscale a bit to help OCR on small screenshots
            w, h = img.size
            if w < 900:
                img = img.resize((int(w * 2), int(h * 2)))
            txt = pytesseract.image_to_string(img)
        except Exception:
            continue

        pv = _extract_int_after_label(txt, r"profile\s+visits")
        elt = _extract_int_after_label(txt, r"external\s+link\s+taps")

        if pv is None and elt is None:
            continue

        rows.append(
            {
                "Account": acct or "",
                "Platform": plat or "",
                "Profile Visits": pv,
                "External Link Taps": elt,
                "SourceFile": fn,
            }
        )

    if not rows:
        return pd.DataFrame(columns=["Account", "Platform", "Profile Visits", "External Link Taps", "SourceFile"])
    return pd.DataFrame(rows)


@st.cache_data
def load_social_data():
    """Legacy loader kept for Sentiment Dashboard compatibility — uses Instagram posts CSV."""
    csv_file = "dataset_instagram-all post-all id.csv"
    if not os.path.exists(csv_file):
        # Fallback to old file if present
        csv_file = "dataset_instagram-scraper_2026-04-10_06-47-02-173.csv"
    if not os.path.exists(csv_file):
        st.error(f"Instagram data file not found.")
        return pd.DataFrame()
    try:
        raw_df = pd.read_csv(csv_file, encoding="utf-8-sig", low_memory=False)
    except Exception as e:
        st.error(f"Error loading CSV: {e}")
        return pd.DataFrame()

    cols_to_keep = ["timestamp", "caption", "likesCount", "commentsCount", "url", "ownerUsername"]
    available_cols = [c for c in cols_to_keep if c in raw_df.columns]

    # Pull hashtag columns (hashtags/0, hashtags/1, …)
    hashtag_cols = [c for c in raw_df.columns if re.match(r"^hashtags/\d+$", c)]

    df = raw_df[available_cols + hashtag_cols].copy()
    if "caption" in df.columns:
        df = df.dropna(subset=["caption"])
    df.rename(columns={"caption": "Post", "timestamp": "Date"}, inplace=True)
    df["Platform"] = "Instagram"
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    likes    = df.get("likesCount",    pd.Series(dtype=float)).fillna(0)
    comments = df.get("commentsCount", pd.Series(dtype=float)).fillna(0)
    df["Engagement"] = likes + comments

    if hashtag_cols:
        df["HashtagList"] = df[hashtag_cols].apply(
            lambda row: tuple(str(h).strip() for h in row if pd.notna(h) and str(h).strip()), axis=1
        )
    else:
        df["HashtagList"] = [tuple() for _ in range(len(df))]

    return df


# ── UNIFIED DATA LOADING (all 5 CSVs) ─────────────────────────────────────────
@st.cache_data
def load_linkedin_data():
    """Load LinkedIn data from xlsx (real hyperlink URLs) + csv (all other columns, as-is)."""
    xlsx_f = "Social Data _ Sep 2025 - March 2026.xlsx"
    csv_f  = "Social Data _ Sep 2025 - March 2026.csv"

    if not os.path.exists(csv_f):
        return pd.DataFrame()

    try:
        try:
            df = pd.read_csv(csv_f, encoding="utf-8-sig", low_memory=False)
            if len(df.columns) == 1 and "\t" in df.columns[0]:
                df = pd.read_csv(csv_f, encoding="utf-8-sig", sep='\t', low_memory=False)
        except UnicodeDecodeError:
            df = pd.read_csv(csv_f, encoding="cp1252", sep='\t', low_memory=False)
            if len(df.columns) == 1 and "," in df.columns[0]:
                df = pd.read_csv(csv_f, encoding="cp1252", sep=',', low_memory=False)
    except Exception:
        return pd.DataFrame()

    # Drop truly empty trailing columns only
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    # ── Extract real hyperlink URLs from the xlsx (openpyxl) ──
    urls = [None] * len(df)
    if os.path.exists(xlsx_f):
        try:
            import openpyxl
            wb  = openpyxl.load_workbook(xlsx_f)
            ws  = wb.active
            # Data rows start at row 2 (row 1 = header)
            for i, row_idx in enumerate(range(2, ws.max_row + 1)):
                cell = ws.cell(row_idx, 1)   # Column A = Post Link
                if cell.hyperlink and hasattr(cell.hyperlink, "target"):
                    urls[i] = cell.hyperlink.target
        except Exception:
            pass

    df["Post Link"] = urls   # replace display-text with real URLs

    # _dt_sort purely for internal sort ordering (never shown to user)
    # Date format in the LinkedIn CSV is DD-MM-YYYY; Time is often missing (NaN),
    # so we parse Date alone first with an explicit format, then add Time when available.
    _date_parsed = pd.to_datetime(df["Date"].astype(str), format="%d-%m-%Y", errors="coerce")
    _time_parsed = pd.to_timedelta(
        df["Time"].astype(str).where(df["Time"].notna() & (df["Time"].astype(str) != "nan"), "0"),
        errors="coerce"
    ).fillna(pd.Timedelta(0))
    df["_dt_sort"] = (_date_parsed + _time_parsed).dt.tz_localize(
        "UTC", ambiguous="NaT", nonexistent="NaT"
    )

    # Compute Engagement for KPI summary only
    def _to_num(col):
        if col in df.columns:
            return pd.to_numeric(
                df[col].astype(str).str.replace(",", "", regex=False), errors="coerce"
            ).fillna(0)
        return pd.Series(0, index=df.index)

    df["Engagement"] = _to_num("Reactions") + _to_num("Comment") + _to_num("Shares")

    df.rename(columns={
        "Message":            "Caption",
        "Click Through Rate": "Click Through Rate %",
    }, inplace=True)

    df["HashtagList"] = [tuple() for _ in range(len(df))]
    df["Source"]      = "LinkedIn"
    df["Account"]     = "HaystackAnalytics"
    df["Platform"]    = "LinkedIn"

    if "Post Type" not in df.columns:
        df["Post Type"] = "Post"

    return df


@st.cache_data
def load_instagram_posts():
    """Load and clean the Instagram all-posts CSV (images + carousels + videos)."""
    # Only our 3 owned Instagram accounts
    OUR_ACCOUNTS = {"haystack_analytics", "infexn_in", "identifi.health"}

    f = "dataset_instagram-all post-all id.csv"
    if not os.path.exists(f):
        return pd.DataFrame()
    try:
        df = pd.read_csv(f, encoding="utf-8-sig", low_memory=False)
    except Exception:
        return pd.DataFrame()

    hashtag_cols = [c for c in df.columns if re.match(r"^hashtags/\d+$", c)]
    keep = ["caption", "ownerUsername", "ownerId", "timestamp",
            "likesCount", "commentsCount", "url", "type", "shortCode"] + hashtag_cols
    df = df[[c for c in keep if c in df.columns]].copy()
    df = df.dropna(subset=["caption"])

    # ── Keep only our owned accounts ──
    if "ownerUsername" in df.columns:
        df = df[df["ownerUsername"].isin(OUR_ACCOUNTS)]

    df["_dt"]     = pd.to_datetime(df["timestamp"], errors="coerce", utc=True)
    df["Date"]    = df["_dt"].dt.strftime("%d/%m/%Y")
    df["Time"]    = df["_dt"].dt.strftime("%H:%M")
    df["_dt_sort"] = df["_dt"]

    df["Reactions"] = pd.to_numeric(df.get("likesCount"), errors="coerce").fillna(0)
    df["Comment"]   = pd.to_numeric(df.get("commentsCount"), errors="coerce").fillna(0)
    df["Engagement"] = df["Reactions"] + df["Comment"]

    # Merge hashtags
    if hashtag_cols:
        df["HashtagList"] = df[hashtag_cols].apply(
            lambda row: tuple(str(h).strip() for h in row if pd.notna(h) and str(h).strip()), axis=1
        )
    else:
        df["HashtagList"] = [tuple() for _ in range(len(df))]

    df.rename(columns={"caption": "Caption", "url": "Post Link",
                        "ownerUsername": "Account", "type": "Post Type"}, inplace=True)

    df["Shares"]             = pd.NA
    df["Clicks"]             = pd.NA
    df["Impressions"]        = pd.NA
    df["Engagement Rate %"]  = pd.NA
    df["Click Through Rate %"] = pd.NA
    df["New Followers from posts"] = pd.NA
    df["Source"]   = "Instagram"
    df["Platform"] = "Instagram"

    return df


@st.cache_data
def load_instagram_reels():
    """Load and clean the Instagram reels CSV."""
    # Only our 3 owned Instagram accounts
    OUR_ACCOUNTS = {"haystack_analytics", "infexn_in", "identifi.health"}

    f = "dataset_instagram-reel all id.csv"
    if not os.path.exists(f):
        return pd.DataFrame()
    try:
        df = pd.read_csv(f, encoding="utf-8-sig", low_memory=False)
    except Exception:
        return pd.DataFrame()

    hashtag_cols = [c for c in df.columns if re.match(r"^hashtags/\d+$", c)]
    keep = ["caption", "ownerUsername", "ownerId", "timestamp",
            "likesCount", "commentsCount", "videoViewCount", "videoPlayCount",
            "url", "type", "shortCode"] + hashtag_cols
    df = df[[c for c in keep if c in df.columns]].copy()
    df = df.dropna(subset=["caption"])

    # ── Keep only our owned accounts ──
    if "ownerUsername" in df.columns:
        df = df[df["ownerUsername"].isin(OUR_ACCOUNTS)]

    df["_dt"]     = pd.to_datetime(df["timestamp"], errors="coerce", utc=True)
    df["Date"]    = df["_dt"].dt.strftime("%d/%m/%Y")
    df["Time"]    = df["_dt"].dt.strftime("%H:%M")
    df["_dt_sort"] = df["_dt"]

    df["Reactions"] = pd.to_numeric(df.get("likesCount"), errors="coerce").fillna(0)
    df["Comment"]   = pd.to_numeric(df.get("commentsCount"), errors="coerce").fillna(0)
    v_view = pd.to_numeric(df.get("videoViewCount"), errors="coerce").fillna(0)
    v_play = pd.to_numeric(df.get("videoPlayCount"), errors="coerce").fillna(0)
    df["Impressions"] = pd.concat([v_view, v_play], axis=1).max(axis=1)
    df["Engagement"]  = df["Reactions"] + df["Comment"]

    if hashtag_cols:
        df["HashtagList"] = df[hashtag_cols].apply(
            lambda row: tuple(str(h).strip() for h in row if pd.notna(h) and str(h).strip()), axis=1
        )
    else:
        df["HashtagList"] = [tuple() for _ in range(len(df))]

    df.rename(columns={"caption": "Caption", "url": "Post Link",
                        "ownerUsername": "Account", "type": "Post Type"}, inplace=True)

    df["Shares"]             = pd.NA
    df["Clicks"]             = pd.NA
    df["Engagement Rate %"]  = pd.NA
    df["Click Through Rate %"] = pd.NA
    df["New Followers from posts"] = pd.NA
    df["Source"]   = "Instagram"
    df["Platform"] = "Instagram"

    return df


@st.cache_data
def load_unified_data():
    """Merge LinkedIn + Instagram Posts + Instagram Reels into one clean DataFrame."""
    UNIFIED_COLS = [
        "Post Link", "Date", "Time", "Caption", "Reactions", "Comment",
        "Shares", "Clicks", "Impressions", "Engagement Rate %",
        "Click Through Rate %", "New Followers from posts",
        "Post Type", "Source", "Account", "Platform",
        "Engagement", "HashtagList", "_dt_sort"
    ]

    frames = []
    for fn in [load_linkedin_data, load_instagram_posts, load_instagram_reels]:
        df = fn()
        if not df.empty:
            # Add missing columns as NA
            for col in UNIFIED_COLS:
                if col not in df.columns:
                    df[col] = pd.NA
            frames.append(df[UNIFIED_COLS])

    if not frames:
        return pd.DataFrame(columns=UNIFIED_COLS)

    unified = pd.concat(frames, ignore_index=True)

    # Sort: LinkedIn first, then all Instagram collated by Account → Post Type → newest first
    PLATFORM_ORDER = {"LinkedIn": 0, "Instagram": 1}
    unified["_platform_order"] = unified["Source"].map(PLATFORM_ORDER).fillna(1).astype(int)

    # Within Instagram: sort by Account (alpha), then Post Type order, then newest first
    TYPE_SUB = {"Image": 0, "Sidecar": 1, "Carousel": 2, "Video": 3}
    unified["_source_sub"] = unified["Post Type"].map(TYPE_SUB).fillna(4).astype(int)

    unified.sort_values(
        ["_platform_order", "Account", "_source_sub", "_dt_sort"],
        ascending=[True, True, True, False],
        inplace=True,
        na_position="last"
    )
    unified.reset_index(drop=True, inplace=True)
    return unified

# ── SENTIMENT ANALYSIS ────────────────────────────────────────────────────────
@st.cache_data
def analyze_sentiment(dataframe):
    if dataframe.empty:
        return dataframe
    results = []
    for text in dataframe["Post"].astype(str).tolist():
        try:
            polarity = TextBlob(text).sentiment.polarity
            results.append("Positive" if polarity > 0.1 else ("Negative" if polarity < -0.1 else "Neutral"))
        except Exception:
            results.append("Neutral")
    dataframe["Sentiment"] = results
    return dataframe

# ── LIVE TRENDS: GNEWS ────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def fetch_gnews(keyword: str, max_results: int = 8):
    """Fetch medical news via GNews RSS (no API key required)."""
    try:
        from gnews import GNews
        gn = GNews(language="en", country="US", max_results=max_results)
        articles = gn.get_news(keyword)
        results = []
        for a in articles:
            results.append({
                "title":       a.get("title", "No title"),
                "description": a.get("description", ""),
                "url":         a.get("url", ""),
                "published":   a.get("published date", ""),
                "source":      a.get("publisher", {}).get("title", "Google News"),
            })
        return results
    except Exception as e:
        return [{"title": f"GNews error: {e}", "description": "", "url": "", "published": "", "source": "Error"}]

# ── HASHTAG SUGGESTION ENGINE ─────────────────────────────────────────────────
def suggest_hashtags(topic_query: str, instagram_df: pd.DataFrame, top_n: int = 20):
    """
    Smart hashtag suggestion:
      1. Curated bank (keyword match)
      2. Mining actual hashtags from the Instagram dataset
      3. Always appends universal general tags
    """
    query_lower = topic_query.lower()
    selected = set()

    # 1️⃣ Curated bank match
    for key, tags in HASHTAG_BANK.items():
        if key.lower() in query_lower or query_lower in key.lower():
            selected.update(tags)

    # Always add general
    selected.update(HASHTAG_BANK["General"])

    # 2️⃣ Mine top hashtags from filtered Instagram data
    if not instagram_df.empty and "HashtagList" in instagram_df.columns:
        # Filter posts that mention the query
        mask = instagram_df["Post"].str.contains(topic_query, case=False, na=False, regex=False)
        relevant = instagram_df[mask]
        if not relevant.empty:
            all_tags = []
            for tag_list in relevant["HashtagList"]:
                all_tags.extend(tag_list)
            # Normalise and count
            tag_counts = {}
            for t in all_tags:
                t_norm = t if t.startswith("#") else f"#{t}"
                tag_counts[t_norm] = tag_counts.get(t_norm, 0) + 1
            # Pick top mined tags
            top_mined = sorted(tag_counts, key=tag_counts.get, reverse=True)[:15]
            selected.update(top_mined)

    result = sorted(selected)[:top_n]
    # Ensure # prefix
    return [t if t.startswith("#") else f"#{t}" for t in result if t and t != "#"]

# ── CLINICAL TERM RANK SEARCH ──────────────────────────────────────────────────
def rank_clinical_terms(instagram_df: pd.DataFrame, search_term: str):
    """
    Given a search term, show:
     - post count / total engagement
     - Google Trends rank (if available)
     - Top posts using that term from Instagram
    """
    if instagram_df.empty or not search_term:
        return pd.DataFrame(), pd.DataFrame()

    mask = instagram_df["Post"].str.contains(search_term, case=False, na=False, regex=False)
    matched = instagram_df[mask].copy()
    if matched.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Aggregate by date
    if "Date" in matched.columns:
        timeline = matched.groupby("Date").agg(
            Mentions=("Post", "count"),
            Total_Engagement=("Engagement", "sum")
        ).reset_index().sort_values("Date")
    else:
        timeline = pd.DataFrame()

    # Top posts
    top_posts_cols = ["Date", "Sentiment", "Engagement", "Post", "url"]
    avail = [c for c in top_posts_cols if c in matched.columns]
    top_posts = matched[avail].sort_values("Engagement", ascending=False).head(10)

    return timeline, top_posts


# ── CONTENT CATEGORIZATION ────────────────────────────────────────────────────
@st.cache_data
def analyze_content_type(dataframe):
    if dataframe.empty:
        return dataframe
    
    def categorize(text):
        text = str(text).lower()
        counts = {
            "Happy/Celebratory": sum(1 for w in ["thrilled", "excited", "happy", "proud", "celebrating", "congratulations", "success", "milestone", "award", "win", "glad", "honored", "welcoming", "joy", "fantastic", "amazing", "grateful", "thank you"] if w in text),
            "Informative/Educational": sum(1 for w in ["did you know", "learn", "how to", "guide", "tips", "discover", "understanding", "study", "research", "science", "facts", "informative", "didyouknow", "insight", "analysis", "information", "report", "clinical", "diagnostic", "data"] if w in text),
            "Promotional": sum(1 for w in ["buy", "discount", "offer", "launch", "new product", "service", "available now", "get yours", "sign up", "register", "sale", "promo", "pricing", "book now", "solution"] if w in text),
            "Interactive/Engaging": sum(1 for w in ["what do you think", "comment below", "share your thoughts", "poll", "question of the day", "tell us", "let us know", "drop a comment", "join the conversation"] if w in text),
            "News/Updates": sum(1 for w in ["update", "announcing", "news", "latest", "event", "upcoming", "join us", "webinar", "press release", "announcement", "live", "conference", "booth"] if w in text)
        }
        max_cat = max(counts, key=counts.get)
        if counts[max_cat] > 0:
            return max_cat
        return "General/Other"

    dataframe["Content Type"] = dataframe["Post"].apply(categorize)
    return dataframe

@st.cache_data
def load_competitor_data():
    """Load the competitor data."""
    f = "dataset_instagram-scraper_comeptitor.csv"
    if not os.path.exists(f):
        return pd.DataFrame()
    try:
        df = pd.read_csv(f, encoding="utf-8-sig", low_memory=False)
    except Exception:
        return pd.DataFrame()

    keep = ["caption", "ownerUsername", "ownerId", "timestamp",
            "likesCount", "commentsCount", "videoViewCount", "videoPlayCount",
            "url", "type", "shortCode"]
    df = df[[c for c in keep if c in df.columns]].copy()
    df = df.dropna(subset=["caption"])

    df["_dt"]     = pd.to_datetime(df["timestamp"], errors="coerce", utc=True)
    df["Date"]    = df["_dt"].dt.strftime("%d/%m/%Y")
    df["Time"]    = df["_dt"].dt.strftime("%H:%M")
    df["_dt_sort"] = df["_dt"]

    df["Reactions"] = pd.to_numeric(df.get("likesCount"), errors="coerce").fillna(0)
    df["Comment"]   = pd.to_numeric(df.get("commentsCount"), errors="coerce").fillna(0)
    v_view = pd.to_numeric(df.get("videoViewCount"), errors="coerce").fillna(0)
    v_play = pd.to_numeric(df.get("videoPlayCount"), errors="coerce").fillna(0)
    
    # Calculate impressions as max of views or plays (if video), otherwise engagement
    if "videoViewCount" in df.columns or "videoPlayCount" in df.columns:
        df["Impressions"] = pd.concat([v_view, v_play], axis=1).max(axis=1)
    else:
        df["Impressions"] = df["Reactions"] + df["Comment"]

    df["Engagement"] = df["Reactions"] + df["Comment"]
    df["HashtagList"] = [tuple() for _ in range(len(df))]

    df.rename(columns={"caption": "Caption", "url": "Post Link",
                        "ownerUsername": "Account", "type": "Post Type"}, inplace=True)

    df["Shares"]             = pd.NA
    df["Clicks"]             = pd.NA
    df["Engagement Rate %"]  = pd.NA
    df["Click Through Rate %"] = pd.NA
    df["New Followers from posts"] = pd.NA
    df["Source"]   = "Competitor"
    df["Platform"] = "Instagram"
    
    # Replace zero impressions with engagement to avoid 0 errors
    df["Impressions"] = df.apply(lambda r: r["Engagement"] if (pd.isna(r["Impressions"]) or r["Impressions"] == 0) else r["Impressions"], axis=1)

    return df

@st.cache_data
def load_similar_accounts_data():
    """Load the similar accounts data."""
    f = "dataset_ofinstagram account that are similar.csv"
    if not os.path.exists(f):
        return pd.DataFrame()
    try:
        df = pd.read_csv(f, encoding="utf-8-sig", low_memory=False)
    except Exception:
        return pd.DataFrame()

    keep = ["caption", "ownerUsername", "ownerId", "timestamp",
            "likesCount", "commentsCount", "videoViewCount", "videoPlayCount",
            "url", "type", "shortCode"]
    df = df[[c for c in keep if c in df.columns]].copy()
    df = df.dropna(subset=["caption"])

    df["_dt"]     = pd.to_datetime(df["timestamp"], errors="coerce", utc=True)
    df["Date"]    = df["_dt"].dt.strftime("%d/%m/%Y")
    df["Time"]    = df["_dt"].dt.strftime("%H:%M")
    df["_dt_sort"] = df["_dt"]

    df["Reactions"] = pd.to_numeric(df.get("likesCount"), errors="coerce").fillna(0)
    df["Comment"]   = pd.to_numeric(df.get("commentsCount"), errors="coerce").fillna(0)
    v_view = pd.to_numeric(df.get("videoViewCount"), errors="coerce").fillna(0)
    v_play = pd.to_numeric(df.get("videoPlayCount"), errors="coerce").fillna(0)
    
    # Calculate impressions as max of views or plays (if video), otherwise engagement
    if "videoViewCount" in df.columns or "videoPlayCount" in df.columns:
        df["Impressions"] = pd.concat([v_view, v_play], axis=1).max(axis=1)
    else:
        df["Impressions"] = df["Reactions"] + df["Comment"]

    df["Engagement"] = df["Reactions"] + df["Comment"]
    df["HashtagList"] = [tuple() for _ in range(len(df))]

    df.rename(columns={"caption": "Caption", "url": "Post Link",
                        "ownerUsername": "Account", "type": "Post Type"}, inplace=True)

    df["Shares"]             = pd.NA
    df["Clicks"]             = pd.NA
    df["Engagement Rate %"]  = pd.NA
    df["Click Through Rate %"] = pd.NA
    df["New Followers from posts"] = pd.NA
    df["Source"]   = "Similar Accounts"
    df["Platform"] = "Instagram"
    
    # Replace zero impressions with engagement to avoid 0 errors
    df["Impressions"] = df.apply(lambda r: r["Engagement"] if (pd.isna(r["Impressions"]) or r["Impressions"] == 0) else r["Impressions"], axis=1)

    return df


# ═══════════════════════════════════════════════════════════════════════════════
# LOAD & PROCESS DATA
# ═══════════════════════════════════════════════════════════════════════════════
df = load_unified_data()
acct_metrics_df = load_account_metrics()
img_metrics_df = load_profile_visit_images()
if not df.empty:
    df["Post"] = df.get("Caption", "")
    df["url"] = df.get("Post Link", "")
    with st.spinner("Analyzing sentiments & content types…"):
        processed_df = analyze_sentiment(df.copy())
        processed_df = analyze_content_type(processed_df)
else:
    processed_df = df

comp_df = load_competitor_data()
if not comp_df.empty:
    comp_df["Post"] = comp_df.get("Caption", "")
    comp_df["url"] = comp_df.get("Post Link", "")
    with st.spinner("Analyzing competitor content..."):
        p_comp_df = analyze_sentiment(comp_df.copy())
        p_comp_df = analyze_content_type(p_comp_df)
else:
    p_comp_df = pd.DataFrame()

sim_df = load_similar_accounts_data()
if not sim_df.empty:
    sim_df["Post"] = sim_df.get("Caption", "")
    sim_df["url"] = sim_df.get("Post Link", "")
    with st.spinner("Analyzing similar accounts content..."):
        p_sim_df = analyze_sentiment(sim_df.copy())
        p_sim_df = analyze_content_type(p_sim_df)
else:
    p_sim_df = pd.DataFrame()

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
# Reddit credentials placeholders
reddit_id       = ""
reddit_secret   = ""


# ═══════════════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style="background: linear-gradient(135deg,#1a1a2e,#16213e);
            border-left: 5px solid #6c63ff; border-radius: 12px;
            padding: 20px 28px; margin-bottom: 24px;">
  <h1 style="color:#e0e0ff; margin:0; font-size:28px;"> Social Media Analytics</h1>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TABS
# ═══════════════════════════════════════════════════════════════════════════════
tab_overall, tab_comp, tab_similar, tab_interp, tab1, tab2, tab3 = st.tabs([
    "Overall Performance",
    "Competitor Benchmarking",
    "Similar Content Accounts",
    "Interpretation & Strategy",
    "Sentiment Dashboard",
    "Clinical & Industry Trends",
    "Cleaned Data Explorer",
])



# ╔══════════════════════════════════════════════════════╗
# ║  TAB 0 — OVERALL PERFORMANCE                         ║
# ╚══════════════════════════════════════════════════════╝
with tab_overall:
    st.markdown('<div class="section-header">Overall Page Health & Performance</div>', unsafe_allow_html=True)

    import numpy as np

    # ── ACCOUNT CONFIG ──────────────────────────────────────────────────────────
    ACCOUNTS = [
        {"label": "Instagram – infexn_in",         "account": "infexn_in",           "platform": "Instagram", "color": "#E1306C"},
        {"label": "Instagram – haystack_analytics", "account": "haystack_analytics",  "platform": "Instagram", "color": "#F77737"},
        {"label": "Instagram – identifi.health",    "account": "identifi.health",     "platform": "Instagram", "color": "#FCAF45"},
        {"label": "LinkedIn – HaystackAnalytics",   "account": "HaystackAnalytics",   "platform": "LinkedIn",  "color": "#0A66C2"},
    ]

    # ── FILTER ROW ──────────────────────────────────────────────────────────────
    ov_c1, ov_c2 = st.columns(2)
    with ov_c1:
        account_options = ["All Accounts"] + [a["label"] for a in ACCOUNTS]
        selected_account_label = st.selectbox("Select Account to Analyze:", account_options, index=0, key="ov_acct")

    # Apply date range filter on full data first
    base_df = processed_df.copy()
    with ov_c2:
        ov_date_range = None
        if not base_df.empty and "_dt_sort" in base_df.columns:
            ov_valid_dt = base_df["_dt_sort"].dropna()
            if not ov_valid_dt.empty:
                ov_min_d = ov_valid_dt.min().date()
                ov_max_d = ov_valid_dt.max().date()
                if pd.notnull(ov_min_d) and pd.notnull(ov_max_d) and ov_min_d != ov_max_d:
                    try:
                        if "ov_date" not in st.session_state:
                            st.session_state["ov_date"] = (ov_min_d, ov_max_d)
                        ov_date_range = st.date_input("Date range:", value=st.session_state["ov_date"],
                                                      min_value=ov_min_d, max_value=ov_max_d, key="ov_date")
                    except Exception:
                        pass

    if ov_date_range and len(ov_date_range) == 2:
        ov_s, ov_e = ov_date_range
        if "_dt_sort" in base_df.columns:
            _s = base_df["_dt_sort"].dt.date
            base_df = base_df[(_s >= ov_s) & (_s <= ov_e)]

        # Filter account metrics on the same date window (date only)
        _m = acct_metrics_df.copy()
        if not _m.empty and "Date" in _m.columns:
            _m = _m[(_m["Date"] >= ov_s) & (_m["Date"] <= ov_e)]
        acct_metrics_f = _m
    else:
        acct_metrics_f = acct_metrics_df.copy()

    # ── HELPER: compute KPIs for a sub-dataframe ────────────────────────────────
    def _calc_kpis(df):
        """Returns df with _CalcImpressions and _CalcFollowers added."""
        if df.empty:
            df["_CalcImpressions"] = pd.Series(dtype=float)
            df["_CalcFollowers"]   = pd.Series(dtype=float)
            return df
        imps, fols = [], []
        for _, row in df.iterrows():
            plat = row.get("Platform", "")
            if plat == "LinkedIn":
                v = row.get("Impressions", 0)
                imps.append(int(pd.to_numeric(str(v).replace(',', ''), errors='coerce') or 0) if pd.notna(v) else 0)
                f = row.get("New Followers from posts", np.nan)
                fols.append(int(pd.to_numeric(str(f).replace(',', ''), errors='coerce') or 0) if pd.notna(f) else np.nan)
            else:
                v = row.get("Impressions", 0)
                if pd.isna(v) or v == 0: v = row.get("Engagement", 0)
                imps.append(int(pd.to_numeric(v, errors='coerce') or 0) if pd.notna(v) else 0)
                fols.append(np.nan)
        df = df.copy()
        df["_CalcImpressions"] = imps
        df["_CalcFollowers"]   = fols
        return df

    # ══════════════════════════════════════════════════════════════════════════
    # DETAILED VIEW (filtered by selected account)
    # ══════════════════════════════════════════════════════════════════════════
    # Build p_df based on account selection
    if selected_account_label == "All Accounts":
        p_df = base_df.copy()
        m_df = acct_metrics_f.copy()
    else:
        sel = next((a for a in ACCOUNTS if a["label"] == selected_account_label), None)
        if sel:
            p_df = base_df[
                (base_df["Account"] == sel["account"]) &
                (base_df["Platform"] == sel["platform"])
            ].copy()
            m_df = acct_metrics_f.copy()
            if not m_df.empty:
                m_df = m_df[
                    (m_df["Account"] == sel["account"]) &
                    (m_df["Platform"] == sel["platform"])
                ].copy()
        else:
            p_df = base_df.copy()
            m_df = acct_metrics_f.copy()

    p_df = _calc_kpis(p_df)

    if p_df.empty:
        st.warning("No data available for the selected account / date range.")
    else:
        # 1. TOP KPI ROW ────────────────────────────────────────────────────────
        c1, c2, c3, c4 = st.columns(4)
        no_posts = len(p_df)
        c1.metric("No of Posts", f"{no_posts:,}")

        ttl_impressions = p_df["_CalcImpressions"].sum()
        c2.metric("Total Impressions / Views", f"{int(ttl_impressions):,}")

        # Followers gain preference:
        # - Use account-level metrics if available (covers Instagram follower-gained files)
        # - Otherwise fall back to LinkedIn "New Followers from posts" column
        fol_from_metrics = None
        if not m_df.empty and "Followers Gained" in m_df.columns:
            _v = pd.to_numeric(m_df["Followers Gained"], errors="coerce")
            if _v.notna().any():
                fol_from_metrics = int(_v.fillna(0).sum())
        # Fallback: if date-range filtering removed metrics, use full metrics for the selected account
        if fol_from_metrics is None and selected_account_label != "All Accounts":
            sel = next((a for a in ACCOUNTS if a["label"] == selected_account_label), None)
            if sel and acct_metrics_df is not None and not acct_metrics_df.empty:
                _all = acct_metrics_df.copy()
                _all = _all[(_all["Account"] == sel["account"]) & (_all["Platform"] == sel["platform"])]
                if not _all.empty and "Followers Gained" in _all.columns:
                    _v2 = pd.to_numeric(_all["Followers Gained"], errors="coerce")
                    if _v2.notna().any():
                        fol_from_metrics = int(_v2.fillna(0).sum())

        if fol_from_metrics is not None:
            c3.metric("Followers Gain", f"{fol_from_metrics:,}")
        elif p_df["_CalcFollowers"].notna().any():
            c3.metric("Followers Gain", f"{int(p_df['_CalcFollowers'].sum()):,}")
        else:
            c3.metric("Followers Gain", "N/A")

        ttl_eng = p_df["Engagement"].sum() if "Engagement" in p_df.columns else 0
        avg_er  = round((ttl_eng / ttl_impressions * 100), 2) if ttl_impressions > 0 else 0
        c4.metric("Avg Engagement Rate %", f"{avg_er}%")

        st.markdown("---")

        # 2. MIDDLE CHARTS ──────────────────────────────────────────────────────
        col_m1, col_m2, col_m3 = st.columns(3)

        # IMPORTANT: Use `_dt_sort` for all time bucketing to avoid DD/MM vs MM/DD mis-parsing.
        # `Date` is a display string in mixed formats, so re-parsing it can create bogus months.
        if "_dt_sort" in p_df.columns:
            _bucket_dt = pd.to_datetime(p_df["_dt_sort"], errors="coerce", utc=True)
        else:
            _bucket_dt = pd.to_datetime(p_df.get("Date", pd.Series(dtype=str)), errors="coerce", dayfirst=True, utc=True)

        if not _bucket_dt.isna().all():
            p_df["Month"]     = _bucket_dt.dt.strftime('%b %Y')
            p_df["DayOfWeek"] = _bucket_dt.dt.day_name()
            cats = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
            p_df["DayOfWeek"] = pd.Categorical(p_df["DayOfWeek"], categories=cats, ordered=True)

            with col_m1:
                st.markdown("**Impressions Every Month**")
                m_df = p_df.groupby("Month").agg({"_CalcImpressions":"sum","Platform":"count"}).rename(columns={"Platform":"Post Count"}).reset_index()
                if not m_df.empty:
                    m_df["sort_date"] = pd.to_datetime(m_df["Month"], format='%b %Y')
                    m_df = m_df.sort_values("sort_date")
                    fig1 = go.Figure()
                    fig1.add_trace(go.Bar(x=m_df["Month"], y=m_df["_CalcImpressions"], name="Impressions", marker_color="#007bff"))
                    fig1.add_trace(go.Scatter(x=m_df["Month"], y=m_df["Post Count"], name="Posts", mode="lines+markers", yaxis="y2", line=dict(color="#172242", width=2)))
                    fig1.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                       yaxis2=dict(overlaying="y", side="right", showgrid=False),
                                       margin=dict(l=0,r=0,t=10,b=0), showlegend=False)
                    st.plotly_chart(fig1, use_container_width=True)

            with col_m2:
                st.markdown("**Impressions by Day of Week**")
                dow_df = p_df.groupby("DayOfWeek")["_CalcImpressions"].sum().reset_index()
                if not dow_df.empty:
                    fig2 = px.line(dow_df, x="DayOfWeek", y="_CalcImpressions", markers=True)
                    fig2.update_traces(line_color="#007bff", line_width=2)
                    fig2.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                       margin=dict(l=0,r=0,t=10,b=0), xaxis_title=None, yaxis_title=None)
                    st.plotly_chart(fig2, use_container_width=True)

            with col_m3:
                st.markdown("**Impressions by Post Type**")
                pt_df = p_df.groupby("Post Type")["_CalcImpressions"].sum().reset_index().sort_values("_CalcImpressions", ascending=False)
                if not pt_df.empty:
                    fig3 = px.bar(pt_df, x="Post Type", y="_CalcImpressions")
                    fig3.update_traces(marker_color="#007bff")
                    fig3.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                       margin=dict(l=0,r=0,t=10,b=0), xaxis_title=None, yaxis_title=None)
                    st.plotly_chart(fig3, use_container_width=True)

            st.markdown("---")

            co1, co2 = st.columns([1.5, 1])
            with co1:
                st.markdown("**Day of Week Breakdown**")
                table_df = p_df.groupby("DayOfWeek").agg({"_CalcImpressions":"sum", "Platform":"count"}).rename(columns={"Platform":"No of Posts"})
                table_df["Click Through Rate"] = "N/A"
                if "Clicks" in p_df.columns:
                    p_df["_CalcClicks"] = pd.to_numeric(p_df["Clicks"].astype(str).str.replace(',',''), errors='coerce').fillna(0)
                    ctr_agg = p_df.groupby("DayOfWeek")["_CalcClicks"].sum()
                    table_df["Click Through Rate"] = round((ctr_agg / table_df["_CalcImpressions"].replace(0,1)) * 100, 2).astype(str) + "%"
                table_df["New Followers"] = p_df.groupby("DayOfWeek")["_CalcFollowers"].sum(min_count=1)
                table_df["New Followers"] = table_df["New Followers"].fillna("N/A")
                st.dataframe(table_df.reset_index(), use_container_width=True)

            with co2:
                st.markdown("**Sentiment Analysis per Month**")
                if "Sentiment" in p_df.columns:
                    sent_df = p_df.groupby(["Month","Sentiment"]).size().reset_index(name="Count")
                    sent_df["sort_date"] = pd.to_datetime(sent_df["Month"], format='%b %Y')
                    sent_df = sent_df.sort_values("sort_date")
                    fig4 = px.bar(sent_df, x="Month", y="Count", color="Sentiment", barmode="stack",
                                  color_discrete_map={"Positive":"#4ecdc4","Neutral":"#6c63ff","Negative":"#ff6b4a"})
                    fig4.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                       margin=dict(l=0,r=0,t=10,b=0), xaxis_title=None)
                    st.plotly_chart(fig4, use_container_width=True)

            st.markdown("---")
            
            st.markdown("**Performance by Content Type**")
            st.caption("Categorization based on the caption: Informative, Happy/Celebratory, etc.")
            if "Content Type" in p_df.columns:
                ct_df = p_df.groupby("Content Type").agg(
                    Posts=("Platform", "count"),
                    Total_Impressions=("_CalcImpressions", "sum")
                ).reset_index()
                # Compute Avg Impressions per Post
                ct_df["Avg_Impressions"] = (ct_df["Total_Impressions"] / ct_df["Posts"]).round(1)
                
                c_ct1, c_ct2 = st.columns(2)
                with c_ct1:
                    if not ct_df.empty:
                        fig_ct = px.pie(ct_df, values='Posts', names='Content Type', hole=0.4, title='Share of Voice by Content Type')
                        fig_ct.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(l=0,r=0,t=30,b=0))
                        st.plotly_chart(fig_ct, use_container_width=True)
                with c_ct2:
                    if not ct_df.empty:
                        fig_ct2 = px.bar(ct_df.sort_values("Avg_Impressions", ascending=False), x="Content Type", y="Avg_Impressions", title="Avg Impressions by Content Type")
                        fig_ct2.update_traces(marker_color="#4ecdc4")
                        fig_ct2.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(l=0,r=0,t=30,b=0), xaxis_title=None)
                        st.plotly_chart(fig_ct2, use_container_width=True)

            st.markdown("---")

            co3, co4 = st.columns([1, 1])
            with co3:
                st.markdown("**Top Performing Past Video**")
                v_df = p_df[p_df["Post Type"].str.lower().isin(["video","reel"])]
                if not v_df.empty:
                    top_v = v_df.sort_values("_CalcImpressions", ascending=False).iloc[0]
                    url = top_v.get("url", top_v.get("Post Link",""))
                    st.success(f"**Views/Impressions:** {int(top_v['_CalcImpressions']):,}\n\n**Date:** {top_v['Date']}\n\n[Click to view video/reel]({url})")
                else:
                    st.info("No videos or reels found in the dataset for this selection.")

                st.markdown("**Hashtags Performance**")
                if "HashtagList" in p_df.columns:
                    h_dict = {}
                    for _, row in p_df.iterrows():
                        _imp = row.get("_CalcImpressions", 0)
                        for h in row.get("HashtagList", []):
                            h_dict[h] = h_dict.get(h, 0) + _imp
                    top_h = sorted(h_dict.items(), key=lambda x: x[1], reverse=True)[:5]
                    if top_h:
                        st.table(pd.DataFrame(top_h, columns=["Hashtag","Total Impressions"]))
                    else:
                        st.caption("N/A")

            with co4:
                st.markdown("**Overall Narration of the Pages**")
                ov_text = f"Analyzed {no_posts} posts generating {int(ttl_impressions):,} impressions. "
                if avg_er > 5:
                    ov_text += f"Engagement rate is highly healthy at {avg_er}%. "
                elif avg_er > 0:
                    ov_text += f"Maintained a stable engagement rate of {avg_er}%. "
                if not p_df.empty and "Sentiment" in p_df.columns:
                    top_sent = p_df["Sentiment"].mode()[0] if not p_df["Sentiment"].mode().empty else "Neutral"
                    ov_text += f"The overarching community sentiment is intensely **{top_sent}**. "
                st.info(ov_text)

                st.markdown("**Other Metrics (Not present in dataset)**")
                prof_visits_val = None
                if not m_df.empty and "Profile Visits" in m_df.columns:
                    _pv = pd.to_numeric(m_df["Profile Visits"], errors="coerce")
                    if _pv.notna().any():
                        prof_visits_val = int(_pv.fillna(0).sum())

                # Per your instruction: fixed profile-visit totals for Instagram accounts
                if selected_account_label != "All Accounts":
                    sel = next((a for a in ACCOUNTS if a["label"] == selected_account_label), None)
                    if sel and sel["platform"] == "Instagram":
                        fixed_pv = {
                            "infexn_in": 522,
                            "identifi.health": 477,
                            "haystack_analytics": 3422,
                        }
                        if sel["account"] in fixed_pv:
                            prof_visits_val = fixed_pv[sel["account"]]

                # If profile visits not found in excel metrics, try screenshot OCR (or manual fallback UI below)
                if prof_visits_val is None and not img_metrics_df.empty and selected_account_label != "All Accounts":
                    sel = next((a for a in ACCOUNTS if a["label"] == selected_account_label), None)
                    if sel:
                        _im = img_metrics_df.copy()
                        _im = _im[(_im["Account"] == sel["account"]) & (_im["Platform"] == sel["platform"])]
                        if not _im.empty and _im["Profile Visits"].notna().any():
                            prof_visits_val = int(pd.to_numeric(_im["Profile Visits"], errors="coerce").fillna(0).sum())

                st.text(f"Profile Visit               : {prof_visits_val:,}" if prof_visits_val is not None else "Profile Visit               : N/A")
                st.text("Video Retention (0.3)       : N/A")
                st.text("Demographics Breakdown      : N/A")

                # Manual fallback: show screenshot and allow entering Profile Visits if OCR isn't available
                if prof_visits_val is None and selected_account_label != "All Accounts":
                    sel = next((a for a in ACCOUNTS if a["label"] == selected_account_label), None)
                    if sel:
                        # find matching image by filename heuristics
                        wanted = sel["account"].lower().split("_")[0]
                        imgs = [f for f in os.listdir(".") if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))]
                        match = None
                        for f in imgs:
                            lf = f.lower()
                            if wanted in lf and ("profile" in lf or "insta" in lf or "instagram" in lf or "linkedin" in lf):
                                match = f
                                break
                        if match:
                            st.image(match, caption=f"Screenshot source: {match}", use_container_width=False)
                        manual_pv = st.number_input(
                            "Manual: Profile Visits (use if not in Excel / OCR unavailable)",
                            min_value=0,
                            value=0,
                            step=1,
                            key=f"manual_pv_{sel['account']}_{sel['platform']}",
                        )
                        if manual_pv and manual_pv > 0:
                            st.caption("Manual value entered (used for this session only).")
                            st.text(f"Profile Visit (manual)      : {int(manual_pv):,}")



# ╔══════════════════════════════════════════════════════╗
# ║  TAB COMP — COMPETITOR BENCHMARKING                  ║
# ╚══════════════════════════════════════════════════════╝
with tab_comp:
    st.markdown('<div class="section-header">Competitor Benchmarking (Us vs. MedGenome)</div>', unsafe_allow_html=True)
    if p_comp_df.empty:
        st.warning("No competitor data available.")
    else:
        # Calculate combined stats for Us vs Them (Instagram Only)
        us_df = processed_df[processed_df["Platform"] == "Instagram"].copy()
        them_df = p_comp_df.copy()

        us_posts = len(us_df)
        them_posts = len(them_df)

        # Helpers for robust summing and impressions handling
        def sum_col(df, col):
            if col in df.columns:
                return pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0).sum()
            return 0
        
        def safe_impressions(df):
            if "Impressions" in df.columns:
                return pd.to_numeric(df["Impressions"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
            return pd.to_numeric(df.get("Engagement", 0), errors='coerce').fillna(0)

        us_eng = sum_col(us_df, "Engagement")
        them_eng = sum_col(them_df, "Engagement")

        us_avg_eng = round(us_eng / us_posts, 1) if us_posts > 0 else 0
        them_avg_eng = round(them_eng / them_posts, 1) if them_posts > 0 else 0
        
        us_imps = safe_impressions(us_df).sum()
        them_imps = safe_impressions(them_df).sum()

        st.subheader("1. Head-to-Head Topline Metrics")
        
        c_kpi1, c_kpi2, c_kpi3 = st.columns(3)
        c_kpi1.metric("Total Posts", f"{us_posts} (Us)", delta=f"{them_posts} (Them)", delta_color="off")
        c_kpi2.metric("Total Engagement", f"{int(us_eng):,} (Us)", delta=f"{int(them_eng):,} (Them)", delta_color="off")
        c_kpi3.metric("Avg Engagement/Post", f"{us_avg_eng} (Us)", delta=f"{them_avg_eng} (Them)", delta_color="off")

        st.markdown("---")
        
        col_c1, col_c2 = st.columns(2)
        
        with col_c1:
            st.subheader("2. Content Strategy Matchup")
            st.caption("What type of content is being produced?")
            if "Content Type" in us_df.columns and "Content Type" in them_df.columns:
                us_ct = us_df["Content Type"].value_counts(normalize=True).reset_index()
                us_ct.columns = ["Content Type", "Percentage"]
                us_ct["Account"] = "Us (Internal)"
                
                them_ct = them_df["Content Type"].value_counts(normalize=True).reset_index()
                them_ct.columns = ["Content Type", "Percentage"]
                them_ct["Account"] = "MedGenome"
                
                combined_ct = pd.concat([us_ct, them_ct])
                combined_ct["Percentage"] = combined_ct["Percentage"] * 100
                
                fig_c1 = px.bar(combined_ct, x="Account", y="Percentage", color="Content Type", barmode="stack")
                fig_c1.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(l=0,r=0,t=10,b=0))
                st.plotly_chart(fig_c1, use_container_width=True)

        with col_c2:
            st.subheader("3. Sentiment & Community Reaction")
            st.caption("How positive or negative are the underlying captions?")
            if "Sentiment" in us_df.columns and "Sentiment" in them_df.columns:
                us_sen = us_df["Sentiment"].value_counts(normalize=True).reset_index()
                us_sen.columns = ["Sentiment", "Percentage"]
                us_sen["Account"] = "Us (Internal)"
                
                them_sen = them_df["Sentiment"].value_counts(normalize=True).reset_index()
                them_sen.columns = ["Sentiment", "Percentage"]
                them_sen["Account"] = "MedGenome"
                
                combined_sen = pd.concat([us_sen, them_sen])
                combined_sen["Percentage"] = combined_sen["Percentage"] * 100
                
                fig_c2 = px.bar(combined_sen, x="Account", y="Percentage", color="Sentiment", barmode="group",
                                color_discrete_map={"Positive":"#4ecdc4","Neutral":"#6c63ff","Negative":"#ff6b4a"})
                fig_c2.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(l=0,r=0,t=10,b=0))
                st.plotly_chart(fig_c2, use_container_width=True)
                
        st.markdown("---")
        st.subheader("4. Top Performing Competitor Posts")
        st.caption("The most engaging content from MedGenome. Use this to draw inspiration.")
        # Fill missing Engagement with Reactions to sort safely
        them_df["_SortEng"] = pd.to_numeric(them_df["Engagement"], errors='coerce').fillna(0)
        top_them = them_df.sort_values(by="_SortEng", ascending=False).head(5)
        for _, row in top_them.iterrows():
            eng = int(pd.to_numeric(row.get("Engagement", 0), errors='coerce'))
            date_val = row.get("Date", "Unknown")
            post_url = row.get("url", "")
            post_txt = str(row.get("Caption", ""))[:400]
            st.markdown(f"""
            <div class="post-card" style="border-left: 4px solid #FCAF45;">
                <div class="post-stats">📅 {{date_val}} · ❤️ {{eng:,}} engagement</div>
                <div class="post-caption">{{post_txt}} ... </div>
                <div style="margin-top:10px;"><a href="{{post_url}}" target="_blank" style="color:#6c63ff;font-size:12px;">View post ↗</a></div>
            </div>
            """.format(date_val=date_val, eng=eng, post_txt=post_txt, post_url=post_url), unsafe_allow_html=True)
            
# ╔══════════════════════════════════════════════════════╗
# ║  TAB SIMILAR — SIMILAR ACCOUNTS ANALYSIS             ║
# ╚══════════════════════════════════════════════════════╝
with tab_similar:
    st.markdown('<div class="section-header">Similar Accounts Analysis</div>', unsafe_allow_html=True)
    if p_sim_df.empty:
        st.warning("No similar accounts data available.")
    else:
        st.write("Evaluating posting patterns of similar Instagram accounts to see what drives their growth.")
        
        c1, c2 = st.columns(2)
        with c1:
            st.subheader("1. Content Type Breakdown")
            st.caption("Which type of content are they posting most frequently?")
            if "Content Type" in p_sim_df.columns:
                ct_counts = p_sim_df["Content Type"].value_counts().reset_index()
                ct_counts.columns = ["Content Type", "Count"]
                fig_ct = px.pie(ct_counts, values="Count", names="Content Type", hole=0.4)
                fig_ct.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_ct, use_container_width=True)
                
        with c2:
            st.subheader("2. Post Type Breakdown")
            st.caption("Which format (Image, Video, Carousel, etc.) is used most?")
            if "Post Type" in p_sim_df.columns:
                pt_counts = p_sim_df["Post Type"].value_counts().reset_index()
                pt_counts.columns = ["Post Type", "Count"]
                fig_pt = px.pie(pt_counts, values="Count", names="Post Type", hole=0.4)
                fig_pt.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig_pt, use_container_width=True)

        st.markdown("---")
        
        st.subheader("3. Engagement by Content & Post Type")
        st.caption("How both metrics contribute to their channel's growth (measured by Impressions/Engagement per post).")
        
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            if "Content Type" in p_sim_df.columns:
                ct_perf = p_sim_df.groupby("Content Type").agg({"Impressions": "mean"}).reset_index()
                ct_perf.columns = ["Content Type", "Avg Impressions"]
                fig_perf_ct = px.bar(ct_perf.sort_values("Avg Impressions", ascending=False), x="Content Type", y="Avg Impressions", color="Content Type")
                fig_perf_ct.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", showlegend=False)
                st.plotly_chart(fig_perf_ct, use_container_width=True)
                
        with col_m2:
            if "Post Type" in p_sim_df.columns:
                pt_perf = p_sim_df.groupby("Post Type").agg({"Impressions": "mean"}).reset_index()
                pt_perf.columns = ["Post Type", "Avg Impressions"]
                fig_perf_pt = px.bar(pt_perf.sort_values("Avg Impressions", ascending=False), x="Post Type", y="Avg Impressions", color="Post Type")
                fig_perf_pt.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", showlegend=False)
                st.plotly_chart(fig_perf_pt, use_container_width=True)
                
        st.markdown("---")
        
        st.subheader("4. Posting Frequency per Account")
        st.caption("Average number of posts per month for each account.")
        if "Date" in p_sim_df.columns and "Account" in p_sim_df.columns:
            freq_df = p_sim_df.dropna(subset=["_dt", "Account"]).copy()
            if not freq_df.empty:
                acct_stats = freq_df.groupby("Account").agg(
                    Total_Posts=("Account", "count"),
                    First_Post=("_dt", "min"),
                    Last_Post=("_dt", "max")
                ).reset_index()
                
                acct_stats["Duration_Days"] = (acct_stats["Last_Post"] - acct_stats["First_Post"]).dt.days
                acct_stats["Duration_Days"] = acct_stats["Duration_Days"].apply(lambda x: max(x, 1))
                acct_stats["Posts_per_Month"] = (acct_stats["Total_Posts"] / (acct_stats["Duration_Days"] / 30.44)).round(1)
                acct_stats = acct_stats.sort_values("Posts_per_Month", ascending=False)
                
                fig_freq = px.bar(acct_stats, x="Account", y="Posts_per_Month", text="Posts_per_Month",
                                  title="Posting Frequency (Posts / Month)")
                fig_freq.update_traces(textposition='outside')
                fig_freq.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", xaxis_tickangle=-45)
                st.plotly_chart(fig_freq, use_container_width=True)
                
                acct_stats["First_Post"] = acct_stats["First_Post"].dt.date
                acct_stats["Last_Post"] = acct_stats["Last_Post"].dt.date
                st.dataframe(acct_stats[["Account", "Total_Posts", "First_Post", "Last_Post", "Posts_per_Month"]], use_container_width=True)
            else:
                st.info("Not enough date information to calculate posting frequency.")

# ╔══════════════════════════════════════════════════════╗
# ║  TAB INTERP — INTERPRETATION & STRATEGY              ║
# ╚══════════════════════════════════════════════════════╝
with tab_interp:
    st.markdown('<div class="section-header">Interpretation & Strategy Insights</div>', unsafe_allow_html=True)
    if processed_df.empty:
        st.info("No data available for interpretation.")
    else:
        # Pre-process numeric representations
        _i_df = processed_df.copy()
        
        # We need numeric Imps across both platforms for logic
        if "Impressions" in _i_df.columns:
            _i_df["_IntImps"] = pd.to_numeric(_i_df["Impressions"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        elif "videoViewCount" in _i_df.columns:
            _i_df["_IntImps"] = pd.to_numeric(_i_df["videoViewCount"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        else:
            _i_df["_IntImps"] = 0
            
        if "Reactions" in _i_df.columns:
            _i_df["_IntReact"] = pd.to_numeric(_i_df["Reactions"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        elif "likesCount" in _i_df.columns:
            _i_df["_IntReact"] = pd.to_numeric(_i_df["likesCount"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        else:
            _i_df["_IntReact"] = 0
            
        if "Shares" in _i_df.columns:
            _i_df["_IntShares"] = pd.to_numeric(_i_df["Shares"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        else:
            _i_df["_IntShares"] = 0
            
        if "Engagement Rate %" in _i_df.columns:
            _i_df["_IntEngR"] = _i_df["Engagement Rate %"].astype(str).str.replace('%', '')
            _i_df["_IntEngR"] = pd.to_numeric(_i_df["_IntEngR"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        else:
            _i_df["_IntEngR"] = 0
            
        if "Click Through Rate %" in _i_df.columns:
            _i_df["_IntCTR"] = _i_df["Click Through Rate %"].astype(str).str.replace('%', '')
            _i_df["_IntCTR"] = pd.to_numeric(_i_df["_IntCTR"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        else:
            _i_df["_IntCTR"] = 0

        # --- 1. Optimal Posting Window Heatmap ---
        st.subheader("1. Optimal Posting Window Heatmap")
        st.caption("Cross-references Day of the Week with Hour of Day against Engagement Rate to find the best time to post.")
        if "_dt_sort" in _i_df.columns:
            _i_df["_Hour"] = _i_df["_dt_sort"].dt.hour
            _i_df["_Day"] = _i_df["_dt_sort"].dt.day_name()
            hp_df = _i_df.dropna(subset=["_Day", "_Hour"])
            if not hp_df.empty:
                heatmap_data = hp_df.groupby(["_Day", "_Hour"])["_IntReact"].mean().unstack(fill_value=0)
                # Reorder days
                days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                heatmap_data = heatmap_data.reindex([d for d in days_order if d in heatmap_data.index])
                # Format to nice integer Heatmap
                st.dataframe(heatmap_data.style.background_gradient(cmap='Blues', axis=None).format("{:.0f}"), use_container_width=True)
            else:
                st.info("Not enough timestamp data for Heatmap.")
        else:
            st.info("No Timestamp column found.")

        st.markdown("---")
        
        c_i1, c_i2 = st.columns(2)
        
        # --- 2. Virality vs Vanity (Shares vs Reactions) ---
        with c_i1:
            st.subheader("2. Virality vs. Vanity Quotient")
            st.caption("Ratio of Shares (Endorsements) to Reactions (Vanity).")
            total_shares = _i_df["_IntShares"].sum()
            total_reacts = _i_df["_IntReact"].sum()
            if total_reacts > 0 and total_shares > 0:
                share_ratio = total_shares / total_reacts
                st.metric("Share-to-Like Ratio", f"{share_ratio:.2f} Shares per Like")
                if share_ratio > 0.1:
                    st.success("High Virality! Your audience trusts your content enough to endorse it to their specific networks.")
                else:
                    st.warning("Low Virality. People are 'liking' but not endorsing. Consider adding a stark Call-To-Action to 'Share with a colleague'.")
            else:
                st.info("Insufficient Share/Reaction data (Instagram missing native shares).")

        # --- 3. Stop-the-Scroll Factor (Hooks Analysis) ---
        with c_i2:
            st.subheader("3. 'Stop-the-Scroll' Hook Analysis")
            st.caption("Highest Click-Through Rate (CTR) captions that hooked the audience.")
            hook_df = _i_df[(_i_df["_IntCTR"] > 0) & (_i_df["Caption"].notna()) & (_i_df["Caption"] != "")]
            if not hook_df.empty:
                top_hooks = hook_df.sort_values("_IntCTR", ascending=False).head(3)
                for _, row in top_hooks.iterrows():
                    trunc_hook = str(row["Caption"])[:100] + "..."
                    st.markdown(f"**CTR {row['_IntCTR']:.1f}%:** _{trunc_hook}_")
            else:
                st.info("Not enough CTR% data available.")

        st.markdown("---")

        c_i3, c_i4 = st.columns(2)
        
        # --- 4. Caption Length Sweet-Spot ---
        with c_i3:
            st.subheader("4. Caption Length vs Conversion")
            st.caption("Identifies the optimal caption length for Engagement.")
            # Calculate word count manually to avoid lambda issues if column is empty
            def count_words(x):
                return len(str(x).split())
            _i_df["_WordCount"] = _i_df["Caption"].apply(count_words)
            def bucket_words(x):
                if x < 20: return "Short (<20 w)"
                elif x <= 75: return "Medium (20-75 w)"
                else: return "Long (75+ w)"
            _i_df["_WCBucket"] = _i_df["_WordCount"].apply(bucket_words)
            bucket_eng = _i_df.groupby("_WCBucket")["_IntEngR"].mean().reset_index()
            if not bucket_eng.empty:
                import plotly.express as px
                fig_wc = px.bar(bucket_eng, x="_WCBucket", y="_IntEngR", 
                                color="_WCBucket", color_discrete_sequence=[_ACCENT, _ACCENT2, "#a0a4b8"])
                fig_wc.update_layout(xaxis_title="", yaxis_title="Avg Engagement Rate %", margin=dict(l=0, r=0, t=10, b=0), height=300, showlegend=False)
                st.plotly_chart(fig_wc, use_container_width=True)
            else:
                st.info("No text data available.")

        # --- 5. The Silent Lurker Metric ---
        with c_i4:
            st.subheader("5. The 'Silent Lurker' B2B Metric")
            st.caption("Estimates true Reach versus Visible Engagement actions.")
            total_imps = _i_df["_IntImps"].sum()
            if total_imps > 0 and total_reacts > 0:
                # Add total comments
                if "Comment" in _i_df.columns:
                    _i_df["_IntCom"] = pd.to_numeric(_i_df["Comment"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                elif "commentsCount" in _i_df.columns:
                    _i_df["_IntCom"] = pd.to_numeric(_i_df["commentsCount"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                else:
                    _i_df["_IntCom"] = 0
                total_comments = _i_df["_IntCom"].sum()
                
                total_interactions = total_reacts + total_comments + total_shares
                lurkers = total_imps - total_interactions
                lurker_pct = (lurkers / total_imps * 100) if total_imps > 0 else 0
                
                st.metric("Total Silent Lurkers", f"{int(lurkers):,}")
                st.metric("Lurker Ratio", f"{lurker_pct:.1f}% of Audience")
                st.info("In healthcare B2B, a high lurker ratio is perfectly normal due to corporate privacy. Your true reach is significantly higher than your 'likes'.")
            else:
                st.info("Not enough Impressions vs Interactions data.")


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 1 — SENTIMENT DASHBOARD (existing functionality)║
# ╚══════════════════════════════════════════════════════╝
with tab1:
    st.markdown('<div class="section-header">Sentiment Dashboard</div>', unsafe_allow_html=True)

    # ── Inline filters (moved from sidebar) ───────────────────────────────────
    with st.expander("Filters", expanded=False):
        _f1, _f2, _f3 = st.columns([2, 2, 2])
        with _f1:
            selected_topics = st.multiselect("Medical Topics:", MEDICAL_TOPICS, default=[])
        with _f2:
            custom_search = st.text_input("Custom keyword:", "")
        with _f3:
            if not processed_df.empty and "_dt_sort" in processed_df.columns:
                valid_dt = processed_df["_dt_sort"].dropna()
                if not valid_dt.empty:
                    min_d = valid_dt.min().date()
                    max_d = valid_dt.max().date()
                    if pd.notnull(min_d) and pd.notnull(max_d) and min_d != max_d:
                        try:
                            date_range = st.date_input("Date range:", [min_d, max_d], min_value=min_d, max_value=max_d)
                        except Exception:
                            date_range = None
                    else:
                        date_range = None
                else:
                    date_range = None
            else:
                date_range = None

    # Apply filters
    search_terms = list(selected_topics)
    if custom_search.strip():
        search_terms.append(custom_search.strip())
    filtered_df = processed_df.copy()
    if search_terms:
        # We handle regex escape to safely construct our pattern:
        pattern = "|".join([f"({re.escape(t)})" for t in search_terms])
        filtered_df = filtered_df[filtered_df["Post"].str.contains(pattern, case=False, na=False, regex=True)]
    if date_range and len(date_range) == 2:
        s, e = date_range
        if "_dt_sort" in filtered_df.columns:
            dt_series = filtered_df["_dt_sort"].dt.date
            filtered_df = filtered_df[(dt_series >= s) & (dt_series <= e)]

    st.markdown("---")

    if filtered_df.empty:
        st.info("No data for the selected filters.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        total_mentions  = len(filtered_df)
        total_eng       = int(filtered_df["Engagement"].sum()) if "Engagement" in filtered_df.columns else 0
        avg_eng         = round(total_eng / total_mentions, 1) if total_mentions > 0 else 0
        pos_count       = len(filtered_df[filtered_df["Sentiment"] == "Positive"]) if "Sentiment" in filtered_df.columns else 0
        neg_count       = len(filtered_df[filtered_df["Sentiment"] == "Negative"]) if "Sentiment" in filtered_df.columns else 0

        c1.metric("Total Mentions",              f"{total_mentions:,}")
        c2.metric("Total Engagement",            f"{total_eng:,}")
        c3.metric("Avg Engagement / Post",       avg_eng)
        c4.metric("Pos / Neg Sentiment",         f"{pos_count} / {neg_count}")

        st.markdown("---")
        col_l, col_r = st.columns(2)
        with col_l:
            st.subheader("Share of Sentiment")
            if "Sentiment" in filtered_df.columns:
                sc = filtered_df["Sentiment"].value_counts().reset_index()
                sc.columns = ["Sentiment", "Count"]
                fig = px.pie(sc, values="Count", names="Sentiment", hole=0.5,
                             color="Sentiment",
                             color_discrete_map={"Positive":"#4ecdc4","Neutral":"#6c63ff","Negative":"#ff6b4a"})
                fig.update_layout(margin=dict(t=0,b=0,l=0,r=0), paper_bgcolor="rgba(0,0,0,0)", font_color="#c8c8e8")
                st.plotly_chart(fig, use_container_width=True)

        with col_r:
            st.subheader("Mentions Over Time")
            if "Date" in filtered_df.columns:
                tl = filtered_df.groupby("Date").size().reset_index(name="Mentions")
                fig2 = px.line(tl, x="Date", y="Mentions", markers=True,
                               line_shape="spline", color_discrete_sequence=["#6c63ff"])
                fig2.update_layout(margin=dict(t=0,b=0,l=0,r=0),
                                   paper_bgcolor="rgba(0,0,0,0)",
                                   plot_bgcolor="rgba(0,0,0,0)",
                                   font_color="#c8c8e8")
                fig2.update_xaxes(gridcolor="#2a2a4a")
                fig2.update_yaxes(gridcolor="#2a2a4a")
                st.plotly_chart(fig2, use_container_width=True)

        st.markdown("---")
        st.subheader("Engagement by Sentiment")
        if "Engagement" in filtered_df.columns and "Sentiment" in filtered_df.columns:
            fig3 = px.box(filtered_df, x="Sentiment", y="Engagement", color="Sentiment",
                          color_discrete_map={"Positive":"#4ecdc4","Neutral":"#6c63ff","Negative":"#ff6b4a"},
                          points="all",
                          hover_data=["url"] if "url" in filtered_df.columns else None)
            fig3.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                               font_color="#c8c8e8")
            fig3.update_xaxes(gridcolor="#2a2a4a")
            fig3.update_yaxes(gridcolor="#2a2a4a")
            st.plotly_chart(fig3, use_container_width=True)

        st.markdown("---")
        st.subheader("Top Mentions Feed")
        disp_cols = ["Date","Platform","Sentiment","Engagement","Post"]
        if "url" in filtered_df.columns:
            disp_cols.append("url")
        avail_disp = [c for c in disp_cols if c in filtered_df.columns]
        tbl = filtered_df[avail_disp]
        if "Engagement" in tbl.columns:
            tbl = tbl.sort_values("Engagement", ascending=False)
        st.dataframe(tbl, use_container_width=True)

        st.markdown("---")
        st.subheader("Actionable Growth Strategies")
        if "Engagement" in filtered_df.columns and "Sentiment" in filtered_df.columns:
            se = filtered_df.groupby("Sentiment")["Engagement"].mean().to_dict()
            if se:
                top_s = max(se, key=se.get)
                recs = []
                if top_s == "Positive":
                    recs.append("**Scale Positive Themes:** Your audience engages disproportionately with positive clinical milestones. Double down on sharing diagnostic breakthroughs and fast turnaround achievements.")
                elif top_s == "Negative":
                    recs.append("**Address Clinical Pain Points:** Negative-sentiment posts capture the most engagement. Create empathetic content addressing frustrations (e.g., long NGS wait times, complex workflows) and show how your solutions help.")
                else:
                    recs.append("**Elevate Educational Content:** Objective, neutral discussions are winning. Focus on clear infographics, protocol breakdowns, and unbiased diagnostic guidelines.")
                if selected_topics:
                    recs.append(f"**Feature Spotlight on '{selected_topics[0]}':** Run a dedicated series or AMA on this topic to capture existing momentum.")
                recs.append("**Content Nuance:** Target medical professionals — emphasise specificity, sensitivity, and scalability metrics rather than broad patient-facing messaging.")
                for r in recs:
                    st.info(r)


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 2 — LIVE INDUSTRY TRENDS                        ║
# ╚══════════════════════════════════════════════════════╝
with tab2:
    st.markdown('<div class="section-header">Live Medical Industry Trends</div>', unsafe_allow_html=True)

    # ── Controls ──────────────────────────────────────────────────────────────
    col_ctrl1, col_ctrl2 = st.columns([3, 1])
    with col_ctrl1:
        trend_topic = st.selectbox("Select topic to explore:", MEDICAL_TOPICS, index=0)
    with col_ctrl2:
        news_count = st.slider("News articles:", 3, 15, 8)

    # ── Google News ────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Latest News — Google News</div>', unsafe_allow_html=True)
    with st.spinner("Fetching latest medical news…"):
        news_query  = f"{trend_topic} medical diagnostics"
        news_items  = fetch_gnews(news_query, max_results=news_count)

    if news_items:
        for article in news_items:
            title = article.get("title", "No title")
            desc  = article.get("description", "")
            url   = article.get("url", "")
            pub   = article.get("published", "")
            src   = article.get("source", "")
            if "error" in title.lower():
                st.warning(title)
                continue
            link_html = f'<a href="{url}" target="_blank" style="color:#6c63ff;">Read article</a>' if url else ""
            st.markdown(f"""
<div class="trend-card">
  <span class="trend-source-badge badge-gnews">Google News</span>
  <span class="trend-meta">{src} · {pub}</span>
  <div class="trend-title">{title}</div>
  <div class="trend-meta">{desc[:220]}{'…' if len(desc) > 220 else ''}</div>
  <div style="margin-top:10px;">{link_html}</div>
</div>
""", unsafe_allow_html=True)
    else:
        st.info("No news articles found.")

    st.markdown("---")

    # ── Hashtag Suggestions ────────────────────────────────────────────
    st.markdown('<div class="section-header">Hashtag Suggestions for Better Reach</div>', unsafe_allow_html=True)
    st.caption(f"Curated + mined from your Instagram dataset for topic: **{trend_topic}**")

    suggested = suggest_hashtags(trend_topic, processed_df, top_n=30)

    # Group into categories
    primary_tags   = [t for t in suggested if trend_topic.lower().replace(" ", "") in t.lower().replace(" ", "")]
    secondary_tags = [t for t in suggested if t not in primary_tags]

    col_ht1, col_ht2 = st.columns(2)
    with col_ht1:
        st.markdown("**Primary (Topic-Specific)**")
        pills_html = "".join([f'<span class="hashtag-pill">{t}</span>' for t in primary_tags[:15]])
        st.markdown(pills_html or "_None found_", unsafe_allow_html=True)

    with col_ht2:
        st.markdown("** Secondary (Reach Boosters)**")
        pills_html2 = "".join([f'<span class="hashtag-pill">{t}</span>' for t in secondary_tags[:15]])
        st.markdown(pills_html2 or "_None found_", unsafe_allow_html=True)

    # Copy-ready string
    all_tags_str = " ".join(suggested[:25])
    st.markdown("** Copy-ready hashtag set:**")
    st.code(all_tags_str, language=None)

    st.markdown("---")

    # ── Posts Using These Trends ───────────────────────────────────────
    st.markdown('<div class="section-header">Instagram Posts Using This Trend</div>', unsafe_allow_html=True)
    st.caption(f"Posts from your dataset mentioning **{trend_topic}** · sorted by engagement")

    if not processed_df.empty:
        mask_trend = processed_df["Post"].str.contains(trend_topic, case=False, na=False, regex=False)
        trend_posts = processed_df[mask_trend].sort_values("Engagement", ascending=False).head(10)

        if not trend_posts.empty:
            for _, row in trend_posts.iterrows():
                post_text = str(row.get("Post", ""))[:500]
                date_val  = row.get("Date", "")
                eng       = int(row.get("Engagement", 0))
                sent      = row.get("Sentiment", "Neutral")
                url_val   = row.get("url", "")
                sent_color = {"Positive":"#4ecdc4","Negative":"#ff6b4a","Neutral":"#6c63ff"}.get(sent,"#6c63ff")
                sent_badge = f'<span style="background:{sent_color};color:#fff;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:600;">{sent}</span>'
                user_val  = row.get("ownerUsername", "")
                user_txt  = f"@{user_val}" if user_val else ""
                link_html = f' · <a href="{url_val}" target="_blank" style="color:#6c63ff;">View post ↗</a>' if url_val else ""
                st.markdown(f"""
<div class="post-card">
  <div style="margin-bottom:8px;">{sent_badge}</div>
  <div class="post-caption">{post_text}{"…" if len(str(row.get("Post",""))) > 500 else ""}</div>
  <div class="post-stats">📅 {date_val} · ❤️ {eng:,} engagement · 👤 {user_txt}{link_html}</div>
</div>
""", unsafe_allow_html=True)
        else:
            st.info(f"No Instagram posts in the dataset mention '{trend_topic}'.")
    else:
        st.info("Dataset not loaded.")


# ── CLINICAL TERM RANKING (Merged into Tab 2) ───────────
    st.markdown("---")
    st.markdown('<div class="section-header"> Clinical Term Rank Search</div>', unsafe_allow_html=True)
    st.markdown(
        "Search any clinical term to see how it ranks in your Instagram dataset — "
        "post volume, engagement trends, sentiment breakdown, and top posts."
    )

    term_search = st.text_input("Enter clinical term (e.g. 'EGFR', 'cfDNA', 'WES', 'TMB'):", "")

    if term_search.strip():
        term = term_search.strip()
        timeline_df, top_posts_df = rank_clinical_terms(processed_df, term)

        if timeline_df.empty and top_posts_df.empty:
            st.warning(f"No posts found mentioning **'{term}'** in the dataset.")
        else:
            # ── KPIs ──────────────────────────────────────────────────────
            total_posts   = len(top_posts_df)
            total_eng_t   = int(top_posts_df["Engagement"].sum()) if "Engagement" in top_posts_df.columns else 0
            avg_eng_t     = round(total_eng_t / total_posts, 1) if total_posts > 0 else 0

            k1, k2, k3 = st.columns(3)
            k1.metric("Posts mentioning term",    f"{total_posts:,}")
            k2.metric("Total Engagement",          f"{total_eng_t:,}")
            k3.metric("Avg Engagement / Post",     avg_eng_t)

            st.markdown("---")

            col_ll, col_rr = st.columns(2)

            # ── Volume timeline ────────────────────────────────────────────
            with col_ll:
                st.subheader(" Mention Volume Over Time")
                if not timeline_df.empty:
                    fig_tl = px.bar(timeline_df, x="Date", y="Mentions",
                                    color_discrete_sequence=["#6c63ff"])
                    fig_tl.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                                         plot_bgcolor="rgba(0,0,0,0)",
                                         font_color="#c8c8e8",
                                         margin=dict(t=10,b=10,l=10,r=10))
                    fig_tl.update_xaxes(gridcolor="#2a2a4a")
                    fig_tl.update_yaxes(gridcolor="#2a2a4a")
                    st.plotly_chart(fig_tl, use_container_width=True)
                else:
                    st.info("No date data available.")

            # ── Sentiment breakdown ────────────────────────────────────────
            with col_rr:
                st.subheader("Sentiment Breakdown for Term")
                if not top_posts_df.empty and "Sentiment" in top_posts_df.columns:
                    sb = top_posts_df["Sentiment"].value_counts().reset_index()
                    sb.columns = ["Sentiment", "Count"]
                    fig_sb = px.bar(sb, x="Sentiment", y="Count", color="Sentiment",
                                    color_discrete_map={"Positive":"#4ecdc4","Neutral":"#6c63ff","Negative":"#ff6b4a"},
                                    text="Count")
                    fig_sb.update_traces(textposition="outside")
                    fig_sb.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                                          plot_bgcolor="rgba(0,0,0,0)",
                                          font_color="#c8c8e8",
                                          showlegend=False,
                                          margin=dict(t=10,b=10,l=10,r=10))
                    fig_sb.update_xaxes(gridcolor="#2a2a4a")
                    fig_sb.update_yaxes(gridcolor="#2a2a4a")
                    st.plotly_chart(fig_sb, use_container_width=True)
                else:
                    st.info("No sentiment data.")

            st.markdown("---")

            # ── Google Trends for term ─────────────────────────────────────
            if also_trend:
                st.subheader(f"Google Trends — '{term}'")
                with st.spinner("Checking Google Trends…"):
                    term_iot, _ = fetch_pytrends([term], timeframe="today 3-m")
                if not term_iot.empty and term in term_iot.columns:
                    fig_gtr = px.area(term_iot, x=term_iot.index, y=term,
                                      color_discrete_sequence=["#4ecdc4"])
                    fig_gtr.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                                           plot_bgcolor="rgba(0,0,0,0)",
                                           font_color="#c8c8e8",
                                           margin=dict(t=10,b=10,l=10,r=10))
                    fig_gtr.update_xaxes(gridcolor="#2a2a4a")
                    fig_gtr.update_yaxes(gridcolor="#2a2a4a", title="Search Interest (0-100)")
                    st.plotly_chart(fig_gtr, use_container_width=True)
                    peak_val = int(term_iot[term].max())
                    avg_val  = round(term_iot[term].mean(), 1)
                    st.caption(f"Peak interest: **{peak_val} / 100** · Average: **{avg_val} / 100** (past 3 months)")
                else:
                    st.info("No Google Trends data for this term (may be too niche or rate-limited).")

            st.markdown("---")

            # ── Top Posts ─────────────────────────────────────────────────
            st.subheader(f"Top Posts Mentioning '{term}' (by engagement)")
            if not top_posts_df.empty:
                for rank, (_, row) in enumerate(top_posts_df.iterrows(), start=1):
                    post_text = str(row.get("Post", ""))
                    # Highlight the term
                    highlighted = re.sub(
                        f"({re.escape(term)})",
                        r'<mark style="background:#6c63ff33;color:#a8a8ff;border-radius:4px;padding:1px 4px;">\1</mark>',
                        post_text[:600], flags=re.IGNORECASE
                    )
                    date_val = row.get("Date", "")
                    eng      = int(row.get("Engagement", 0))
                    sent     = row.get("Sentiment", "Neutral")
                    url_val  = row.get("url", "")
                    sent_color = {"Positive":"#4ecdc4","Negative":"#ff6b4a","Neutral":"#6c63ff"}.get(sent,"#6c63ff")
                    link_html  = f'<a href="{url_val}" target="_blank" style="color:#6c63ff;font-size:12px;">View post ↗</a>' if url_val else ""
                    st.markdown(f"""
<div class="post-card">
  <div style="display:flex; align-items:center; gap:10px; margin-bottom:10px;">
    <span class="rank-number">{rank}</span>
    <span style="background:{sent_color};color:#fff;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:600;">{sent}</span>
    <span style="color:#8888aa;font-size:12px;">📅 {date_val} · ❤️ {eng:,} engagement</span>
  </div>
  <div class="post-caption">{highlighted}{"…" if len(post_text) > 600 else ""}</div>
  <div style="margin-top:10px;">{link_html}</div>
</div>
""", unsafe_allow_html=True)
            else:
                st.info("No posts to display.")

            st.markdown("---")

            # ── Hashtag suggestions for searched term ──────────────────────
            st.subheader(f"Hashtag Suggestions for '{term}'")
            term_tags = suggest_hashtags(term, processed_df, top_n=20)
            term_tags_html = "".join([f'<span class="hashtag-pill">{t}</span>' for t in term_tags])
            st.markdown(term_tags_html, unsafe_allow_html=True)
            st.code(" ".join(term_tags), language=None)
    else:
        st.info("Enter a clinical term above to analyse its rank and performance.")


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 3 — CLEANED DATA EXPLORER                       ║
# ╚══════════════════════════════════════════════════════╝
with tab3:
    st.markdown('<div class="section-header"> Cleaned Data Explorer</div>', unsafe_allow_html=True)
    st.markdown(
        "Unified view of all social data — **LinkedIn**, **Instagram Posts** and **Instagram Reels** — "
        "cleaned to a common schema. Use the controls below to filter and sort."
    )

    with st.spinner("Loading & cleaning all data sources…"):
        unified_df = load_unified_data()

    if unified_df.empty:
        st.error("⚠️ No data found. Make sure the CSV files are in the same folder as sentiment.py.")
    else:
        # ── ACCOUNT-LEVEL METRICS (Followers gained, Profile visits) ─────────────
        st.markdown("---")
        st.markdown(
            '<div class="section-header" style="font-size:16px;">Account Metrics</div>',
            unsafe_allow_html=True
        )
        if acct_metrics_df is None or acct_metrics_df.empty:
            st.caption("No account-metrics file detected yet. Drop your follower-gained/profile-visit Excel/CSV into this folder to populate this section.")
        else:
            # Filter metrics by the same account selector and date range as the explorer
            m_expl = acct_metrics_df.copy()
            if "sel_accounts" in locals() and sel_accounts:
                m_expl = m_expl[m_expl["Account"].isin(sel_accounts)]
            # date_from/date_to are defined below; so we defer the filtering until after they exist

        # ── FILTER CONTROLS ───────────────────────────────────────────────
        st.markdown('<div class="section-header" style="font-size:16px;">Filters</div>',
                    unsafe_allow_html=True)
        fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 2])

        with fc1:
            all_sources  = sorted(unified_df["Source"].dropna().unique().tolist())
            sel_sources  = st.multiselect("Source:", all_sources, default=all_sources)

        with fc2:
            all_accounts = sorted(unified_df["Account"].dropna().unique().tolist())
            sel_accounts = st.multiselect("Account:", all_accounts, default=all_accounts)

        with fc3:
            all_types = sorted(unified_df["Post Type"].dropna().unique().tolist())
            sel_types = st.multiselect("Post Type:", all_types, default=all_types,
                                       key="post_type_ms")

        with fc4:
            kw_filter = st.text_input("Keyword in caption:", "")

        # Date range
        dc1, dc2 = st.columns(2)
        valid_dates = unified_df["_dt_sort"].dropna()
        min_date = valid_dates.min().date() if not valid_dates.empty else datetime.date(2024, 1, 1)
        max_date = valid_dates.max().date() if not valid_dates.empty else datetime.date.today()
        with dc1:
            date_from = st.date_input("From:", value=min_date, min_value=min_date, max_value=max_date,
                                       key="explorer_date_from")
        with dc2:
            date_to   = st.date_input("To:",   value=max_date, min_value=min_date, max_value=max_date,
                                       key="explorer_date_to")

        # Now that date_from/date_to are known, filter metrics and show KPIs + table
        if acct_metrics_df is not None and not acct_metrics_df.empty:
            m_expl = acct_metrics_df.copy()
            if sel_accounts:
                m_expl = m_expl[m_expl["Account"].isin(sel_accounts)]
            m_expl = m_expl[(m_expl["Date"] >= date_from) & (m_expl["Date"] <= date_to)]

            mk1, mk2, mk3 = st.columns(3)
            fol_sum = pd.to_numeric(m_expl.get("Followers Gained", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
            pv_sum  = pd.to_numeric(m_expl.get("Profile Visits", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
            mk1.metric("Followers gained (selected)", f"{int(fol_sum):,}")
            mk2.metric("Profile visits (selected)", f"{int(pv_sum):,}")
            mk3.metric("Metric rows", f"{len(m_expl):,}")
            st.dataframe(m_expl.sort_values(["Date", "Account", "Platform"]), use_container_width=True, height=min(360, 45 + 28 * len(m_expl)))

        # ── SORT CONTROLS ─────────────────────────────────────────────────
        st.markdown('<div class="section-header" style="font-size:16px;">↕️ Sort</div>',
                    unsafe_allow_html=True)
        sc1, sc2 = st.columns([3, 1])
        sortable_cols = ["Date", "Reactions", "Comment", "Shares", "Clicks",
                         "Impressions", "Engagement Rate %", "Click Through Rate %",
                         "Engagement", "New Followers from posts"]
        with sc1:
            sort_col = st.selectbox("Sort by:", sortable_cols, index=0)
        with sc2:
            sort_dir = st.radio("Direction:", ["Descending", "Ascending"], index=0, horizontal=True)

        # ── APPLY FILTERS ─────────────────────────────────────────────────
        expl = unified_df.copy()

        if sel_sources:
            expl = expl[expl["Source"].isin(sel_sources)]
        if sel_accounts:
            expl = expl[expl["Account"].isin(sel_accounts)]
        if sel_types:
            expl = expl[expl["Post Type"].isin(sel_types)]
        if kw_filter.strip():
            expl = expl[expl["Caption"].astype(str).str.contains(
                kw_filter.strip(), case=False, na=False, regex=False)]

        # Date filter — convert date_from / date_to to UTC-aware Timestamps for comparison
        if "_dt_sort" in expl.columns:
            from_ = pd.Timestamp(date_from, tz="UTC")
            to_   = pd.Timestamp(date_to,   tz="UTC") + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
            expl = expl[
                (expl["_dt_sort"].isna()) |
                ((expl["_dt_sort"] >= from_) & (expl["_dt_sort"] <= to_))
            ]

        # Sort — LinkedIn first, then Instagram grouped by Account (posts+reels collated)
        ascending = "Ascending" in sort_dir
        if "_platform_order" not in expl.columns:
            PLATFORM_ORDER = {"LinkedIn": 0, "Instagram": 1}
            expl["_platform_order"] = expl["Source"].map(PLATFORM_ORDER).fillna(1).astype(int)
        if "_source_sub" not in expl.columns:
            TYPE_SUB = {"Image": 0, "Sidecar": 1, "Carousel": 2, "Video": 3}
            expl["_source_sub"] = expl["Post Type"].map(TYPE_SUB).fillna(4).astype(int)

        if sort_col == "Date":
            expl = expl.sort_values(
                ["_platform_order", "Account", "_source_sub", "_dt_sort"],
                ascending=[True, True, True, ascending],
                na_position="last"
            )
        else:
            expl = expl.sort_values(
                ["_platform_order", "Account", "_source_sub", sort_col],
                ascending=[True, True, True, ascending],
                na_position="last"
            )

        # ── SUMMARY KPIs ──────────────────────────────────────────────────
        st.markdown("---")
        k1, k2, k3, k4, k5 = st.columns(5)

        def _safe_sum(series):
            """Sum a column that may contain raw string numbers like '1,945'."""
            return pd.to_numeric(
                series.astype(str).str.replace(",", "", regex=False), errors="coerce"
            ).fillna(0).sum()

        k1.metric("Total Posts",       f"{len(expl):,}")
        k2.metric("Total Reactions",   f"{int(_safe_sum(expl['Reactions'])):,}")
        k3.metric("Total Comments",    f"{int(_safe_sum(expl['Comment'])):,}")
        k4.metric("Total Impressions", f"{int(_safe_sum(expl['Impressions'])):,}")
        k5.metric("Avg Engagement",    f"{round(expl['Engagement'].fillna(0).mean(), 1)}")

        st.markdown("---")

        # ── SOURCE DISTRIBUTION MINI-CHART ────────────────────────────────
        ch1, ch2 = st.columns(2)
        with ch1:
            st.subheader("Posts by Source")
            src_counts = expl["Source"].value_counts().reset_index()
            src_counts.columns = ["Source", "Count"]
            fig_src = px.pie(
                src_counts, values="Count", names="Source", hole=0.5,
                color="Source",
                color_discrete_map={
                    "LinkedIn":   "#0a66c2",
                    "Instagram":  "#e1306c",
                }
            )
            fig_src.update_layout(
                margin=dict(t=0, b=0, l=0, r=0),
                paper_bgcolor="rgba(0,0,0,0)",
                font_color="#c8c8e8",
                legend=dict(font=dict(color="#c8c8e8"))
            )
            st.plotly_chart(fig_src, use_container_width=True)

        with ch2:
            st.subheader("Posts by Account")
            acc_counts = expl["Account"].value_counts().reset_index()
            acc_counts.columns = ["Account", "Count"]
            fig_acc = px.bar(
                acc_counts, x="Account", y="Count",
                color="Account",
                color_discrete_sequence=["#6c63ff", "#4ecdc4", "#ff6b4a", "#ffd166", "#06d6a0"]
            )
            fig_acc.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font_color="#c8c8e8",
                showlegend=False,
                margin=dict(t=10, b=10, l=10, r=10)
            )
            fig_acc.update_xaxes(gridcolor="#2a2a4a")
            fig_acc.update_yaxes(gridcolor="#2a2a4a")
            st.plotly_chart(fig_acc, use_container_width=True)

        st.markdown("---")

        # ── STYLED TABLE ──────────────────────────────────────────────────
        st.subheader(f" {len(expl):,} posts matched")

        # ── Shared numeric helpers ────────────────────────────────────────
        def _fmt_int(v):
            if pd.isna(v) or str(v).strip() in ("", "nan", "None", "<NA>"):
                return "--"
            try:
                return f"{int(float(str(v).replace(',', ''))):,}"
            except Exception:
                return str(v)

        def _fmt_pct(v):
            if pd.isna(v) or str(v).strip() in ("", "nan", "None", "<NA>"):
                return "--"
            try:
                return f"{float(str(v).replace(',', '').replace('%', '')):.2f}%"
            except Exception:
                return str(v)

        def _trunc(s, n=140):
            s = str(s)
            return s[:n] + "..." if len(s) > n else s

        # ── LINKEDIN TABLE ────────────────────────────────────────────────
        li_data = expl[expl["Source"] == "LinkedIn"].copy()
        if not li_data.empty:
            st.markdown(
                '<div class="section-header" style="font-size:16px;">'
                'LinkedIn Posts</div>', unsafe_allow_html=True
            )
            li_cols = ["Post Link", "Date", "Time", "Caption",
                       "Reactions", "Comment", "Shares", "Clicks",
                       "Impressions", "Engagement Rate %", "Click Through Rate %",
                       "Post Type", "Account"]
            li_disp = li_data[[c for c in li_cols if c in li_data.columns]].copy()
            li_disp["Caption"] = li_disp["Caption"].apply(_trunc)
            for nc in ["Reactions", "Comment", "Shares", "Clicks", "Impressions"]:
                if nc in li_disp.columns:
                    li_disp[nc] = li_disp[nc].apply(_fmt_int)
            for pc in ["Engagement Rate %", "Click Through Rate %"]:
                if pc in li_disp.columns:
                    li_disp[pc] = li_disp[pc].apply(_fmt_pct)
            st.dataframe(
                li_disp,
                use_container_width=True,
                height=min(460, 45 + 35 * len(li_disp)),
                column_config={
                    "Post Link": st.column_config.LinkColumn("Post Link", display_text="Open"),
                    "Caption":   st.column_config.TextColumn("Message", width="large"),
                }
            )

        # ── INSTAGRAM TABLE ───────────────────────────────────────────────
        ig_data = expl[expl["Source"] == "Instagram"].copy()
        if not ig_data.empty:
            st.markdown(
                '<div class="section-header" style="font-size:16px;">'
                'Instagram Posts & Reels</div>', unsafe_allow_html=True
            )
            # Map internal columns to image schema:
            #   Reactions  → Likes
            #   Post Type  → Type
            #   Impressions → Views  (videoViewCount for reels)
            #   Reach / Followers / Repost → not in scraper, shown as —
            ig_disp = pd.DataFrame()
            ig_disp["Post link/date"] = ig_data["Post Link"]
            ig_disp["Account"]        = ig_data["Account"]
            ig_disp["Date"]           = ig_data["Date"]
            ig_disp["Time"]           = ig_data["Time"]
            ig_disp["Type"]           = ig_data.get("Post Type", pd.Series("--", index=ig_data.index)).values
            ig_disp["Likes"]          = ig_data.get("Reactions",   pd.Series(0, index=ig_data.index)).apply(_fmt_int).values
            ig_disp["Comment"]        = ig_data.get("Comment",     pd.Series(0, index=ig_data.index)).apply(_fmt_int).values
            ig_disp["Reach"]          = "--"
            ig_disp["Views"]          = ig_data.get("Impressions", pd.Series(pd.NA, index=ig_data.index)).apply(_fmt_int).values
            ig_disp["Followers"]      = "--"
            ig_disp["Shares"]         = "--"
            ig_disp["Repost"]         = "--"

            st.dataframe(
                ig_disp,
                use_container_width=True,
                height=min(520, 45 + 35 * len(ig_disp)),
                column_config={
                    "Post link/date": st.column_config.LinkColumn(
                        "Post link/date", display_text="Open"
                    ),
                }
            )

        # ── EXPORT ────────────────────────────────────────────────────────
        st.markdown("---")
        ex1, ex2 = st.columns([1, 5])
        with ex1:
            export_cols = ["Post Link", "Date", "Time", "Caption",
                           "Reactions", "Comment", "Shares", "Clicks",
                           "Impressions", "Engagement Rate %", "Click Through Rate %",
                           "Post Type", "Source", "Account"]
            export_df = expl[[c for c in export_cols if c in expl.columns]].copy()
            csv_bytes = export_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label="Export filtered data as CSV",
                data=csv_bytes,
                file_name=f"cleaned_social_data_{datetime.date.today()}.csv",
                mime="text/csv",
                use_container_width=True,
            )